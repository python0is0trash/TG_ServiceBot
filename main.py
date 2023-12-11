# -*- coding: utf-8 -*-

from telebot import types
import sqlite3
import pandas as pd
import openpyxl
import os
import glob
from datetime import (datetime,
                      timedelta)
import schedule
from threading import Thread
from time import sleep

from config import *
from service_functions import (num_to_emoji,
                               links_from_start,
                               links_without_button_back,
                               error_from_user,
                               delete_reply_markup,
                               info_to_chat,
                               is_right_string,
                               role_from_db_to_text)
from create_db import (create_table_students,
                       create_table_schedule,
                       create_table_users)


create_table_students()
create_table_schedule()
create_table_users()


def schedule_checker():
    while True:
        schedule.run_pending()
        sleep(1)


@bot.message_handler(
    commands=['start', 'logout', 'main_page',
              'schedule', 'help', 'id', 'all', 'add_info', 'update_students', 'update_users',
              'add_schedule', 'current_schedule', 'subscribe_to_schedule',
              'add_cadet', 'update_cadet', 'delete_cadet',
              'update_user', 'delete_user',
              'hospital_phone'])
def all_commands(message):
    chat_id = message.chat.id

    try:
        if users[chat_id].user_role:
            if message.text == '/start':
                delete_reply_markup(message=message)
                return command_start(message=message)
            elif message.text == '/logout':
                delete_reply_markup(message=message)
                return command_logout(message=message)
            elif message.text == '/help':
                delete_reply_markup(message=message,
                                    text_hint='Возвращаю на страницу помощи...')
                return command_help(message=message, do_edit_message=False)
            elif message.text == '/id':
                delete_reply_markup(message=message,
                                    text_hint='Возвращаю на страницу информации о пользователе...')
                return command_my_id(message=message, do_edit_message=False)

            elif users[chat_id].user_role != 'guest':
                if message.text == '/main_page':
                    delete_reply_markup(message=message,
                                        text_hint='Возвращаю на главную страницу...')
                    return command_main_page(message=message)
                elif message.text == '/all':
                    delete_reply_markup(message=message,
                                        text_hint='Возвращаю на страницу информации о курсантах...')
                    return command_all(message=message, do_edit_message=False, process_index=1)
                elif message.text == '/schedule':
                    delete_reply_markup(message=message,
                                        text_hint='Возвращаю на страницу расписания...')
                    return command_schedule(message=message, do_edit_message=False)
                elif message.text == '/current_schedule':
                    delete_reply_markup(message=message,
                                        text_hint='Возвращаю на страницу информации о расписании...')
                    return create_schedule_courses(message, do_edit_message=False, process_index=1)
                elif message.text == '/subscribe_to_schedule':
                    delete_reply_markup(message=message,
                                        text_hint='Возвращаю на страницу подписки на расписание...')
                    return create_schedule_courses(message, do_edit_message=False, process_index=2)
                elif message.text == '/hospital_phone':
                    delete_reply_markup(message=message,
                                        text_hint='Возвращаю на страницу номера телефона...')
                    return command_hospital_phone(message=message)

                elif users[chat_id].user_role != 'cadet':
                    if message.text == '/add_info':
                        delete_reply_markup(message=message,
                                            text_hint='Возвращаю на страницу добавления информации о курсантах...')
                        return command_all(message=message, do_edit_message=False, process_index=2)
                    elif message.text == '/add_schedule':
                        delete_reply_markup(message=message,
                                            text_hint='Возвращаю на страницу добавления расписания...')
                        return schedule_document_rules(message=message)

                    elif users[chat_id].user_role != 'helper':
                        if message.text == '/update_students':
                            delete_reply_markup(message=message,
                                                text_hint='Возвращаю на страницу добавления информации о курсантах...')
                            return update_students(message, do_edit_message=False)
                        elif message.text == '/add_cadet':
                            delete_reply_markup(message=message,
                                                text_hint='Возвращаю на страницу добавления курсанта...')
                            return command_add_cadet(message=message, do_edit_message=False)
                        elif message.text == '/update_cadet':
                            delete_reply_markup(message=message,
                                                text_hint='Возвращаю на страницу обновления информации о курсантах...')
                            return command_all(message=message, do_edit_message=False, process_index=4)
                        elif message.text == '/delete_cadet':
                            delete_reply_markup(message=message,
                                                text_hint='Возвращаю на страницу информации о курсантах...')
                            return command_all(message=message, do_edit_message=False, process_index=3)
                        elif message.text == '/update_users':
                            delete_reply_markup(message=message,
                                                text_hint='Возвращаю на страницу обновления информации о курсантах...')
                            return command_update_users(message, do_edit_message=False)
                        elif message.text == '/update_user':
                            delete_reply_markup(message=message,
                                                text_hint='Возвращаю на страницу обновления информации о пользователях...')
                            return command_update_user(message=message, process_index=1)
                        elif message.text == '/delete_user':
                            delete_reply_markup(message=message,
                                                text_hint='Возвращаю на страницу удаления пользователей...')
                            return command_update_user(message=message, process_index=2)
                else:
                    markup_inline = types.InlineKeyboardMarkup()
                    button_back = types.InlineKeyboardButton('🔙 Назад', callback_data='back_to_main_page')
                    markup_inline.add(button_back)
                    bot.send_message(chat_id=chat_id,
                                     text='У Вас нет прав на использование данной команды!',
                                     reply_markup=markup_inline)
            else:
                markup_inline = types.InlineKeyboardMarkup()
                button_back = types.InlineKeyboardButton('🔙 Назад', callback_data='back_to_start')
                markup_inline.add(button_back)
                bot.send_message(chat_id=chat_id,
                                 text='У Вас нет прав на использование данной команды!',
                                 reply_markup=markup_inline)
    except KeyError:
        users[chat_id] = User()
        users[chat_id].user_id = message.from_user.id
        users[chat_id].user_name = message.from_user.first_name
        users[chat_id].user_role = 'guest'
        return all_commands(message=message)


# стартовое окно выбора метода входа
def command_start(message, do_edit_message=False):
    chat_id = message.chat.id

    if (users[chat_id].user_role == 'officer') or (users[chat_id].user_role == 'helper') or (
            users[chat_id].user_role == 'cadet'):
        return command_logout(message=message)
    else:
        markup_inline = types.InlineKeyboardMarkup()
        markup_inline.add(types.InlineKeyboardButton(text='🚪 Войти', callback_data='login'),
                          types.InlineKeyboardButton(text='▶ Зарегистрироваться', callback_data='signup'))
        markup_inline.add(types.InlineKeyboardButton(text='❓ Помощь', callback_data='from_mainpage_to_help'))
        markup_inline.add(types.InlineKeyboardButton(text='👋 Информация о пользователе',
                                                     callback_data='from_mainpage_to_info_about_user'))

        if do_edit_message:
            bot.edit_message_text(chat_id=chat_id,
                                  message_id=message.message_id,
                                  text=start_message,
                                  reply_markup=markup_inline)
        else:
            bot.send_message(chat_id=chat_id,
                             text=start_message,
                             reply_markup=markup_inline)


# стартовое окно регистрации (окно ввода логина)
def signup_in_system(message):
    chat_id = message.chat.id

    markup_inline = types.InlineKeyboardMarkup()
    markup_inline.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_start'))
    bot.edit_message_text(chat_id=chat_id,
                          message_id=message.message_id,
                          text='▶ Введите логин, который вы будете использовать при авторизации в будущем (без пробелов), '
                               'например \"<b>user1234</b>\".',
                          parse_mode='html',
                          reply_markup=markup_inline)
    bot.register_next_step_handler(message, signup_check_login)


# окно ввода пароля
def signup_check_login(message):
    chat_id = message.chat.id

    try:
        if message.text in command_list:
            return all_commands(message=message)

        elif (message.text) and (is_right_string(message.text)):
            with sqlite3.connect('./data_bases/users.sqlite3') as connection:
                cursor = connection.cursor()
            cursor.execute('''
                SELECT 
                    user_login 
                FROM 
                    users 
                WHERE 
                    user_login == ('%s');
            '''
                           % (str(message.text)))
            login_overlap = cursor.fetchone()
            connection.commit()
            cursor.close()
            connection.close()
            if login_overlap:
                bot.send_message(chat_id=chat_id,
                                 text='🚨 Пользователь с таким логином уже существует. Выберите другой логин.',
                                 reply_markup=links_from_start())
                bot.register_next_step_handler(message, signup_check_login)
            else:
                markup_inline = types.InlineKeyboardMarkup()
                markup_inline.add(types.InlineKeyboardButton(text='🔁 Ввести другой логин', callback_data='signup'))

                bot.send_message(chat_id=chat_id,
                                 text='▶ Теперь введите пароль, который вы будете использовать при авторизации в будущем (без пробелов).',
                                 reply_markup=markup_inline)
                bot.register_next_step_handler(message, signup_check_password, register_login=message.text.strip())
        else:
            bot.send_message(chat_id=chat_id,
                             text='🚨 Введен некорректный логин. Введите, пожалуйста, корректный логин.',
                             reply_markup=links_from_start())
            bot.register_next_step_handler(message, signup_check_login)
    except TypeError:
        bot.send_message(chat_id=chat_id,
                         text='🚨 Введен некорректный логин. Введите, пожалуйста, корректный логин.',
                         reply_markup=links_from_start())
        bot.register_next_step_handler(message, signup_check_login)


# довавление пользователя в базу данных
def signup_check_password(message, register_login=None):
    chat_id = message.chat.id

    try:
        if message.text in command_list:
            return all_commands(message=message)

        elif (message.text) and (is_right_string(message.text)):
            new_user_password = message.text.strip()

            with sqlite3.connect('./data_bases/users.sqlite3') as connection:
                cursor = connection.cursor()
            cursor.execute('''
                INSERT INTO 
                    users (
                        user_login, 
                        user_password) 
                VALUES ('%s', '%s');
            '''
                           % (str(register_login), str(new_user_password)))
            connection.commit()
            cursor.close()
            connection.close()

            bot.send_message(chat_id=chat_id,
                             text=f'🎉 Пользователь <b>{register_login}</b> успешно зарегистрирован! '
                                  f'Ваш пароль: <b>{new_user_password}</b>.\n\n'
                                  f''
                                  f'🆔 Для продолжения работы необходимо войти в систему.',
                             parse_mode='html')
            return login_in_system(message=message, do_edit_message=False)

        else:
            markup_inline = types.InlineKeyboardMarkup()
            markup_inline.row(types.InlineKeyboardButton(text='🔗 Telegram', url=tg_profile_link),
                              types.InlineKeyboardButton(text='🔗 VK', url=vk_profile_link))
            markup_inline.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='signup'))
            bot.send_message(chat_id=chat_id,
                             text='🚨 Введен некорректный пароль. Введите, пожалуйста, корректный пароль.',
                             reply_markup=markup_inline)
            bot.register_next_step_handler(message, signup_check_password, register_login=register_login)

    except TypeError:
        markup_inline = types.InlineKeyboardMarkup()
        markup_inline.row(types.InlineKeyboardButton(text='🔗 Telegram', url=tg_profile_link),
                          types.InlineKeyboardButton(text='🔗 VK', url=vk_profile_link))
        markup_inline.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='signup'))
        bot.send_message(chat_id=chat_id,
                         text='🚨 Введен некорректный пароль. Введите, пожалуйста, корректный пароль.',
                         reply_markup=markup_inline)
        bot.register_next_step_handler(message, signup_check_password, register_login=register_login)


# стартовое окно авторизации (окно ввода логина)
def login_in_system(message, do_edit_message=False):
    chat_id = message.chat.id

    markup_inline = types.InlineKeyboardMarkup()
    markup_inline.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_start'))

    if do_edit_message:
        bot.edit_message_text(chat_id=chat_id,
                              message_id=message.message_id,
                              text='👋 Для начала работы необходимо пройти авторизацию.\n'
                                   'Введите свой логин, указанный при регистрации, например \"<b>user1234</b>\".',
                              parse_mode='html',
                              reply_markup=markup_inline)
    else:
        bot.send_message(chat_id=chat_id,
                         text='👋 Для начала работы необходимо пройти регистрацию.\n'
                              'Введите свой логин, указанный при регистрации, например \"<b>user1234</b>\".',
                         parse_mode='html',
                         reply_markup=markup_inline)
    bot.register_next_step_handler(message, login_check_login)


# окно ввода пароля
def login_check_login(message):
    chat_id = message.chat.id

    try:
        if message.text in command_list:
            return all_commands(message=message)

        elif (message.text) and (is_right_string(message.text)):
            current_login = message.text

            with sqlite3.connect('./data_bases/users.sqlite3') as connection:
                cursor = connection.cursor()
            cursor.execute('''
                SELECT 
                    user_login 
                FROM 
                    users 
                WHERE user_login == ('%s');
            '''
                           % str(current_login))
            logins = cursor.fetchone()
            connection.commit()
            cursor.close()
            connection.close()

            if logins:
                markup_inline = types.InlineKeyboardMarkup()
                markup_inline.add(types.InlineKeyboardButton(text='🔁 Ввести другой логин', callback_data='login'))
                bot.send_message(chat_id=chat_id,
                                 text=f'✔ Рады видеть Вас снова, <b>{current_login}</b>! Теперь введите свой пароль.',
                                 parse_mode='html',
                                 reply_markup=markup_inline)
                bot.register_next_step_handler(message, login_check_password, user_login=current_login)
            else:
                bot.send_message(chat_id=chat_id,
                                 text='❌ Пользователя с таким логином не найдено! '
                                      'Введите, пожалуйста, корректный логин или обратитесь в поддержку.',
                                 parse_mode='html',
                                 reply_markup=links_from_start())
                bot.register_next_step_handler(message, login_check_login)
        else:
            bot.send_message(chat_id=chat_id,
                             text='❌ Введен неверный формат логина! '
                                  'Введите, пожалуйста, корректный логин или обратитесь в поддержку.',
                             parse_mode='html',
                             reply_markup=links_from_start())
            bot.register_next_step_handler(message, login_check_login)

    except TypeError:
        bot.send_message(chat_id=chat_id,
                         text='❌ Введен неверный формат логина! '
                              'Введите, пожалуйста, корректный логин или обратитесь в поддержку.',
                         parse_mode='html',
                         reply_markup=links_from_start())
        bot.register_next_step_handler(message, login_check_login)


# проверка существования пользователя с данным логином и паролем
def login_check_password(message, user_login=None):
    chat_id = message.chat.id

    try:
        if message.text in command_list:
            return all_commands(message=message)

        elif (message.text) and (is_right_string(message.text)):
            current_password = message.text

            with sqlite3.connect('./data_bases/users.sqlite3') as connection:
                cursor = connection.cursor()
            cursor.execute('''
                SELECT 
                    user_id,   
                    user_login, 
                    user_role, 
                    tg_chat_id 
                FROM 
                    users 
                WHERE 
                    user_login == ('%s') AND user_password == ('%s');
            '''
                           % (str(user_login), str(current_password)))
            user_password = cursor.fetchone()

            if user_password:
                if user_password[3] == default_tg_chat_id:
                    cursor.execute('''
                        UPDATE 
                            users 
                        SET 
                            tg_chat_id = ('%d') 
                        WHERE 
                            user_id == ('%d');
                    '''
                                   % (int(chat_id), int(user_password[0]))
                                   )
                else:
                    cursor.execute('''
                        INSERT INTO 
                            users (
                                user_login, 
                                user_password, 
                                user_role, 
                                tg_chat_id) 
                        VALUES ('%s', '%s', '%s', '%d');
                    '''
                                   % (str(user_password[1]), str(current_password), str(user_password[2]), int(chat_id))
                                   )

                users[chat_id] = User()
                users[chat_id].user_id = message.from_user.id
                users[chat_id].user_name = message.from_user.first_name
                users[chat_id].user_role = user_password[2]
                connection.commit()
                cursor.close()
                connection.close()
                msg = bot.send_message(chat_id=chat_id,
                                 text=f'🔓 Вы успешно вошли в систему под логином <u>{user_password[1]}</u>!',
                                 parse_mode='html')
                sleep(1)
                bot.delete_message(chat_id=chat_id,
                                   message_id=msg.message_id)
                return command_main_page(message=message, do_edit_message=False)
            else:
                markup_inline = types.InlineKeyboardMarkup()
                markup_inline.row(types.InlineKeyboardButton(text='🔗 Telegram', url=tg_profile_link),
                                  types.InlineKeyboardButton(text='🔗 VK', url=vk_profile_link))
                markup_inline.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='login'))
                bot.send_message(chat_id=chat_id,
                                 text=f'❌ Пользователя <b>{user_login}</b> с таким паролем не найдено! '
                                      'Введите, пожалуйста, корректный пароль или обратитесь в поддержку.',
                                 parse_mode='html',
                                 reply_markup=markup_inline)
                cursor.close()
                connection.close()
                bot.register_next_step_handler(message, login_check_password, user_login=user_login)
        else:
            markup_inline = types.InlineKeyboardMarkup()
            markup_inline.row(types.InlineKeyboardButton(text='🔗 Telegram', url=tg_profile_link),
                              types.InlineKeyboardButton(text='🔗 VK', url=vk_profile_link))
            markup_inline.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='login'))
            bot.send_message(chat_id=chat_id,
                             text='❌ Введен неверный формат логина! '
                                  'Введите, пожалуйста, корректный логин или обратитесь в поддержку.',
                             parse_mode='html',
                             reply_markup=markup_inline)
            bot.register_next_step_handler(message, login_check_password, user_login=user_login)

    except TypeError:
        markup_inline = types.InlineKeyboardMarkup()
        markup_inline.row(types.InlineKeyboardButton(text='🔗 Telegram', url=tg_profile_link),
                          types.InlineKeyboardButton(text='🔗 VK', url=vk_profile_link))
        markup_inline.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='login'))
        bot.send_message(chat_id=chat_id,
                         text='❌ Введен неверный формат логина! '
                              'Введите, пожалуйста, корректный логин или обратитесь в поддержку.',
                         parse_mode='html',
                         reply_markup=markup_inline)
        bot.register_next_step_handler(message, login_check_password, user_login=user_login)


def command_logout(message):
    chat_id = message.chat.id

    if (users[chat_id].user_role == 'officer') or (users[chat_id].user_role == 'helper') or (
            users[chat_id].user_role == 'cadet'):
        with sqlite3.connect('./data_bases/users.sqlite3') as connection:
            cursor = connection.cursor()
        cursor.execute('''
            SELECT 
                user_login 
            FROM 
                users 
            WHERE 
                tg_chat_id == ('%d');
        '''
                       % (int(chat_id)))
        is_user_login_from_db = cursor.fetchone()
        connection.commit()
        cursor.close()
        connection.close()

        markup_inline = types.InlineKeyboardMarkup()
        markup_inline.add(types.InlineKeyboardButton(text='🔴 Да, выйти из профиля', callback_data='log_out'),
                          types.InlineKeyboardButton(text='🟢 Нет, остаться', callback_data='stay_on_system'))
        bot.send_message(chat_id=chat_id,
                         text=f'Вы действительно хотите выйти из профиля <b>{is_user_login_from_db[0]}</b>?',
                         parse_mode='html',
                         reply_markup=markup_inline)
    else:
        markup_inline = types.InlineKeyboardMarkup()
        markup_inline.add(types.InlineKeyboardButton(text='▶ Да, зарегистрироваться', callback_data='signup'),
                          types.InlineKeyboardButton(text='❌ Нет, остаться', callback_data='back_to_start'))
        bot.send_message(chat_id=chat_id,
                         text=f'❕ Вы не зарегистрированы! Зарегистрироваться? ❕',
                         reply_markup=markup_inline)


def logout_function(message):
    chat_id = message.chat.id

    try:
        with sqlite3.connect('./data_bases/users.sqlite3') as connection:
            cursor = connection.cursor()
        cursor.execute('''
            SELECT 
                user_login 
            FROM 
                users 
            WHERE 
                tg_chat_id == ('%d');
        '''
                       % (int(chat_id)))
        user_login_help = cursor.fetchone()
        cursor.execute('''
            SELECT 
                COUNT(user_login) 
            FROM 
                users 
            WHERE 
                user_login == ('%s');
        '''
                       % (str(user_login_help[0])))
        count = cursor.fetchone()

        if (count[0]) > 1:
            cursor.execute('''
                DELETE FROM 
                    users 
                WHERE 
                    tg_chat_id == ('%d');
            '''
                           % (int(chat_id))
                           )
        else:
            cursor.execute('''
                UPDATE 
                    users 
                SET 
                    tg_chat_id = {default_tg_id} 
                WHERE 
                    tg_chat_id == {tg_id_from_user}; 
            '''
                           .format(default_tg_id=int(default_tg_chat_id),
                                   tg_id_from_user=int(chat_id)))
        connection.commit()
        cursor.close()
        connection.close()
        users[chat_id].user_role = 'guest'

        return command_start(message=message, do_edit_message=True)

    except Exception as e:
        return error_from_user(message=message, error_code=e)


def command_hospital_phone(message):
    bot.send_contact(chat_id=message.chat.id,
                     phone_number=hospital_phone,
                     first_name='ЦП №2')


# отправка текущего расписания при подключении подписки на него
def subscribe_to_schedule(help_course=None, help_platoon=None, help_group=None, chat_id=None):
    # подключение к базе данных
    tomorrow_day = str((datetime.now() + timedelta(days=1)).day)
    if len(tomorrow_day) == 1:
        tomorrow_day = '0' + tomorrow_day
    tomorrow_month = str((datetime.now() + timedelta(days=1)).month)
    if len(tomorrow_month) == 1:
        tomorrow_month = '0' + tomorrow_month

    with sqlite3.connect('./data_bases/schedule.sqlite3') as connection_to_db_main:
        cursor_db = connection_to_db_main.cursor()
    cursor_db.execute('''
        SELECT 
            li.schedule_lesson_time, 
            li.schedule_lesson, 
            li.schedule_type_of_lesson, 
            li.schedule_tutor, 
            li.schedule_lesson_room
        FROM 
            lessons_info as li
        LEFT OUTER JOIN 
            group_info AS gi
        ON 
            li.group_id = gi.group_id
        WHERE 
            li.schedule_lesson_day == ('%s') AND li.schedule_lesson_month == ('%s') AND schedule_course == ('%d') AND schedule_platoon == ('%s') AND schedule_group == ('%s')
        ORDER BY 
            li.schedule_lesson_time;
    '''
                      % (str(tomorrow_day), str(tomorrow_month), int(help_course), str(help_platoon), str(help_group)))
    lessons_tomorrow = cursor_db.fetchall()

    connection_to_db_main.commit()
    cursor_db.close()
    connection_to_db_main.close()

    lessons_count = 1

    info_lessons = (
        f'{days_week_to_emoji[(datetime.now() + timedelta(days=1)).weekday()]} <i><b><u>Расписание на завтра</u></b></i>, {num_to_emoji(str(tomorrow_day))}.{num_to_emoji(str(tomorrow_month))}, {days_week[(datetime.now() + timedelta(days=1)).weekday()]}, '
        f'<i><b><u>для</u></b></i> {num_to_emoji(str(help_course))}<i><b><u>го курса</u></b></i> {num_to_emoji(str(help_platoon).split(" ")[0])}-{num_to_emoji(str(help_group).split("/")[-1])}<i><b><u>го взвода:</u></b></i>\n\n')
    if lessons_tomorrow:
        for current_lesson in lessons_tomorrow:
            info_lessons += (f'{num_to_emoji(str(lessons_count))} 🕰 \t{current_lesson[0]} 🕰\n'
                             f'🎓 \t<b>{current_lesson[1]}</b>\n'
                             f'🔬 \t{current_lesson[2]} / 👨 {current_lesson[3]}\n'
                             f'🚪 \t{current_lesson[4]}\n\n')
            lessons_count += 1
    else:
        info_lessons += '🧧 Нет занятий!\n\n'

    bot.send_message(chat_id=chat_id, text=info_lessons, parse_mode='html')


# отмена отправки текущего расписания и отключение подписки на него
def unsubscribe_to_schedule(message):
    chat_id = message.chat.id
    try:
        schedule.clear()
        user = users[chat_id]
        user.is_subscribe_to_schedule = False

        markup_inline = types.InlineKeyboardMarkup()
        markup_inline.add(types.InlineKeyboardButton(text='✔ Хорошо', callback_data='back_to_schedule'))
        bot.edit_message_text(chat_id=chat_id,
                              message_id=message.message_id,
                              text='Вы <b>отписались</b> от ежедневного уведомления о занятиях.',
                              reply_markup=markup_inline,
                              parse_mode='html')
    except Exception as e:
        return error_from_user(message=message, error_code=e)


def command_main_page(message, do_edit_message=False):
    chat_id = message.chat.id

    markup_inline = types.InlineKeyboardMarkup()
    button_shedule = types.InlineKeyboardButton(text='📅 Расписание занятий', callback_data='from_mainpage_to_schedule')
    button_help = types.InlineKeyboardButton(text='❓ Помощь', callback_data='from_mainpage_to_help')
    button_all = types.InlineKeyboardButton(text='👥 Информация о курсантах', callback_data='from_mainpage_to_all')
    button_id = types.InlineKeyboardButton(text='👋 Информация о пользователе',
                                           callback_data='from_mainpage_to_info_about_user')
    button_add_info = types.InlineKeyboardButton(text='➕ Добавить информацию о курсантах',
                                                 callback_data='from_mainpage_to_add')
    button_update_students = types.InlineKeyboardButton(text='👨‍🎓 Обновить курсанта', callback_data='from_mainpage_to_update_students')
    button_update_users = types.InlineKeyboardButton(text='👨‍💻 Обновить пользователя', callback_data='from_mainpage_to_update_users')

    if (users[chat_id].user_role == 'officer') or (users[chat_id].user_role == 'helper'):
        markup_inline.row(button_shedule)
        markup_inline.row(button_all, button_add_info)
        if users[chat_id].user_role == 'officer':
            markup_inline.row(button_update_students)
            markup_inline.row(button_update_users)
        markup_inline.row(button_help, button_id)
    elif users[chat_id].user_role == 'cadet':
        markup_inline.row(button_shedule)
        markup_inline.row(button_all)
        markup_inline.row(button_help, button_id)

    if do_edit_message:
        bot.edit_message_text(chat_id=chat_id,
                              message_id=message.message_id,
                              text='▶ Для работы с ботом используйте <b>команды</b> (знак \"/\" в поле ввода) или <b>навигационное меню</b> ниже.',
                              reply_markup=markup_inline,
                              parse_mode='html')
    else:
        bot.send_message(chat_id=chat_id,
                         text='▶ Для работы с ботом используйте <b>команды</b> (знак \"/\" в поле ввода) или <b>навигационное меню</b> ниже.',
                         reply_markup=markup_inline,
                         parse_mode='html')


def update_students(message, do_edit_message=False):
    chat_id = message.chat.id

    markup_inline = types.InlineKeyboardMarkup()
    markup_inline.add(types.InlineKeyboardButton(text='👨🎓 Добавить курсанта',
                                                 callback_data='from_updatestudents_to_add_student'))
    markup_inline.add(types.InlineKeyboardButton(text='👨🔁 Обновить курсанта',
                                                 callback_data='from_updatestudents_to_update_student'),
                      types.InlineKeyboardButton(text='👨❌ Удалить курсанта',
                                                 callback_data='from_updatestudents_to_delete_student'))
    markup_inline.add(types.InlineKeyboardButton(text='🏠 На главную',
                                                 callback_data='back_to_main_page'))

    if do_edit_message:
        bot.edit_message_text(chat_id=chat_id,
                              message_id=message.message_id,
                              text='❔ Что именно Вас интересует?',
                              reply_markup=markup_inline,
                              parse_mode='html')
    else:
        bot.send_message(chat_id=chat_id,
                         text='❔ Что именно Вас интересует?',
                         reply_markup=markup_inline,
                         parse_mode='html')


def command_update_users(message, do_edit_message=False):
    chat_id = message.chat.id

    markup_inline = types.InlineKeyboardMarkup()
    markup_inline.add(types.InlineKeyboardButton(text='👨‍💻🔁 Обновить пользователя',
                                                 callback_data='from_updateusers_to_update_user'))
    markup_inline.add(types.InlineKeyboardButton(text='👨‍💻❌ Удалить пользователя',
                                                 callback_data='from_updateusers_to_delete_user'))
    markup_inline.add(types.InlineKeyboardButton(text='🏠 На главную',
                                                 callback_data='back_to_main_page'))

    if do_edit_message:
        bot.edit_message_text(chat_id=chat_id,
                              message_id=message.message_id,
                              text='❔ Что именно Вас интересует?',
                              reply_markup=markup_inline,
                              parse_mode='html')
    else:
        bot.send_message(chat_id=chat_id,
                         text='❔ Что именно Вас интересует?',
                         reply_markup=markup_inline,
                         parse_mode='html')


# обработка вызова помощи
def command_help(message, do_edit_message=False):
    chat_id = message.chat.id

    markup_inline = types.InlineKeyboardMarkup()
    markup_inline.add(types.InlineKeyboardButton(text='🔗 Telegram', url=tg_profile_link))
    markup_inline.add(types.InlineKeyboardButton(text='🔗 VK', url=vk_profile_link))

    if users[chat_id].user_role != 'guest':
        markup_inline.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_main_page'))
    else:
        markup_inline.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_start'))

    if do_edit_message:
        bot.edit_message_text(chat_id=chat_id,
                              message_id=message.message_id,
                              text=help_command_description,
                              reply_markup=markup_inline,
                              parse_mode='html'
                              )
    else:
        bot.reply_to(message=message,
                     text=help_command_description,
                     parse_mode='html',
                     reply_markup=markup_inline)


def command_my_id(message, do_edit_message=False):
    chat_id = message.chat.id

    try:
        markup_inline = types.InlineKeyboardMarkup()
        if users[chat_id].user_role != 'guest':
            markup_inline.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_main_page'))
        else:
            markup_inline.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_start'))

        if do_edit_message:
            bot.edit_message_text(chat_id=chat_id,
                                  message_id=message.message_id,
                                  text=f'Здравствуйте,\n'
                                       f'\t├ <b>{users[chat_id].user_name}</b>\n'
                                       f'Ваш уникальный ID:\n'
                                       f'\t└ <b>{users[chat_id].user_id}</b>.',
                                  reply_markup=markup_inline,
                                  parse_mode='html')
        else:
            bot.send_message(chat_id=chat_id,
                             text=f'Здравствуйте,\n'
                                  f'\t├ <b>{users[chat_id].user_name}</b>\n'
                                  f'Ваш уникальный ID:\n'
                                  f'\t└ <b>{users[chat_id].user_id}</b>.',
                             reply_markup=markup_inline,
                             parse_mode='html')

    except Exception:
        if do_edit_message:
            bot.edit_message_text(chat_id=chat_id,
                                  message_id=message.message_id,
                                  text='Для получения информации перезайдите в учетную запись или обратитесь в поддержку:',
                                  reply_markup=links_without_button_back(),
                                  parse_mode='html')
        else:
            bot.send_message(chat_id=chat_id,
                             text='Для получения информации перезайдите в учетную запись или обратитесь в поддержку:',
                             reply_markup=links_without_button_back(),
                             parse_mode='html')


# меню выбора действия с расписанием
def command_schedule(message, do_edit_message=False):
    chat_id = message.chat.id

    try:
        markup_inline = types.InlineKeyboardMarkup()
        button_add_file = types.InlineKeyboardButton(text='⬇ Добавить файл расписания курса/взвода',
                                                     callback_data='add_file_from_user_callback')
        button_today_schedule = types.InlineKeyboardButton(text='📆 Текущее расписание',
                                                           callback_data='today_schedule_callback')
        if users[chat_id].is_subscribe_to_schedule:
            button_subscribe_to_schedule = types.InlineKeyboardButton(
                text='🚫 Отменить подписку на уведомление о расписании',
                callback_data='unsubscribe_to_schedule_callback')
        else:
            button_subscribe_to_schedule = types.InlineKeyboardButton(text='🔔 Подписаться на уведомление о расписании',
                                                                      callback_data='subscribe_to_schedule_callback')
        button_back = types.InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_main_page')

        if (users[chat_id].user_role == 'officer') or (users[chat_id].user_role == 'helper'):
            markup_inline.add(button_add_file)
        markup_inline.add(button_today_schedule)
        markup_inline.add(button_subscribe_to_schedule)
        markup_inline.add(button_back)
        if do_edit_message:
            bot.edit_message_text(chat_id=chat_id,
                                  message_id=message.message_id,
                                  text='Выберите нужную опцию.',
                                  reply_markup=markup_inline)
        else:
            bot.send_message(chat_id=chat_id,
                             text='Выберите нужную опцию.',
                             reply_markup=markup_inline)

    except Exception as e:
        return error_from_user(message=message, error_code=e)


# стартовое меню загрузки документа с расписанием
def schedule_document_rules(message):
    chat_id = message.chat.id

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                       one_time_keyboard=True,
                                       input_field_placeholder='Пришлите файл или нажмите кнопку \"🔙 Назад\"...')
    markup.add(types.KeyboardButton(text='🔙 Назад'))
    bot.send_photo(chat_id=chat_id,
                   photo=open('./images/schedule_example_img.png', 'rb'),
                   caption=schedule_file_rules,
                   parse_mode='html',
                   reply_markup=markup)
    bot.register_next_step_handler(message, take_document)


# уведомление пользователя о том, что он прислал неверный формат документа с расписанием
def schedule_document_rules_error(message):
    chat_id = message.chat.id

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                       one_time_keyboard=True,
                                       input_field_placeholder='Пришлите файл или нажмите кнопку \"🔙 Назад\"...')
    markup.add(types.KeyboardButton(text='🔙 Назад'))

    bot.send_message(chat_id=chat_id,
                     text='ℹ Выбран неверный формат файла. Пожалуйста, вышлите отформатированный по правилам файл Excel (.xlsx).',
                     reply_markup=markup)
    bot.register_next_step_handler(message, take_document)


# получение первичной информации о присланном документе расписания
def check_schedule_document_from_user(message, msg_help=None):
    chat_id = message.chat.id

    try:
        main_table = openpyxl.load_workbook(users[chat_id].excel_schedule_file_name)
        worksheet = main_table.active

        max_column = worksheet.max_column

        groups_from_document = {}

        # начало чтения информации
        start_row = 1
        while worksheet.cell(row=start_row, column=1).value != 'Дата':
            start_row += 1

        # получение списка взводов в формате {'взвод1': [], 'взвод2': []}
        for current_platoon in range(3, max_column + 1):
            if str(worksheet.cell(row=start_row, column=current_platoon).value) != 'None':
                groups_from_document[str(worksheet.cell(row=start_row, column=current_platoon).value)] = []

        # получение списка групп в формате {'взвод1': ['группа1', 'группа2'], 'взвод2': ['группа1', 'группа2']}
        help_platoon = ''
        count_group = 1
        for current_group in range(3, max_column + 1):
            if str(worksheet.cell(row=start_row + 1, column=current_group).value) != 'None':
                if str(worksheet.cell(row=start_row, column=current_group).value) != 'None':
                    count_group = 1
                    help_platoon = str(worksheet.cell(row=start_row, column=current_group).value)
                    groups_from_document[help_platoon].append(str(worksheet.cell(row=start_row + 1, column=current_group).value))
                    count_group += 1
                else:
                    groups_from_document[help_platoon].append(str(worksheet.cell(row=start_row + 1, column=current_group).value))
                    count_group += 1
            else:
                if str(worksheet.cell(row=start_row, column=current_group).value) != 'None':
                    count_group = 1
                    help_platoon = str(worksheet.cell(row=start_row, column=current_group).value)
                    groups_from_document[help_platoon].append(str(worksheet.cell(row=start_row + 1, column=current_group).value))
                    count_group += 1

        list_info_from_document = ''
        for current_platoon in groups_from_document.keys():
            if current_platoon.split(" ")[0].isdigit():
                list_info_from_document += str(current_platoon) + ' взвод:\nГруппы '
                for current_group in range(len(groups_from_document[str(current_platoon)]) - 1):
                    if groups_from_document[str(current_platoon)][current_group].split("/")[-1].isdigit():
                        list_info_from_document += groups_from_document[str(current_platoon)][current_group] + ', '
                    else:
                        if len(groups_from_document[current_platoon]) <= 1:
                            del groups_from_document[current_platoon]
                        else:
                            del groups_from_document[str(current_platoon)][current_group]
                list_info_from_document += groups_from_document[str(current_platoon)][len(groups_from_document[str(current_platoon)]) - 1] + '.\n\n'
            else:
                del groups_from_document[current_platoon]

        markup_inline = types.InlineKeyboardMarkup()
        markup_inline.row(types.InlineKeyboardButton(text='✔ Сохранить', callback_data='apply_to_save_document'),
                          types.InlineKeyboardButton(text='❌ Отменить', callback_data='deny_to_save_document'))

        users[chat_id].excel_schedule_file_name = users[chat_id].excel_schedule_file_name.split('/')[-1]

        bot.delete_message(chat_id=chat_id,
                           message_id=msg_help.message_id)
        bot.send_message(chat_id=chat_id,
                         text=f'Вы прислали следующую информацию о взводах/группах:\n\n'
                              f''
                              f'{list_info_from_document}'
                              f''
                              f'Желаете сохранить информацию?',
                         reply_markup=markup_inline)
        return groups_from_document

    except Exception as e:
        return error_from_user(message=message, error_code=e)


# описание того, если пользователь согласится отправить данные документа с расписанием
def apply_to_save_schedule_document(message):
    chat_id = message.chat.id

    markup_inline = types.InlineKeyboardMarkup()
    markup_inline.row(types.InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_schedule'),
                      types.InlineKeyboardButton(text='🔁 Повторить отправку',
                                                 callback_data='add_file_from_user_callback'))
    day_yesterday = str((datetime.now() - timedelta(days=1)).day)
    month_yesterday = str((datetime.now() - timedelta(days=1)).month)
    if len(day_yesterday) == 1:
        day_yesterday = '0' + day_yesterday
    if len(month_yesterday) == 1:
        month_yesterday = '0' + month_yesterday
    schedule_document_to_db(message, day_yesterday, month_yesterday)

    try:
        os.remove('./xlsx_files/' + str(users[chat_id].excel_schedule_file_name))

        bot.edit_message_text(chat_id=chat_id,
                              message_id=message.message_id,
                              text=f'Файл \"<b>{users[chat_id].excel_schedule_file_name}</b>\" успешно сохранен!',
                              reply_markup=markup_inline,
                              parse_mode='html')
    except FileNotFoundError:
        files = glob.glob('./xlsx_files/*')
        for file in files:
            os.remove(file)
        bot.edit_message_text(chat_id=chat_id,
                              message_id=message.message_id,
                              text=f'Файл \"<b>{users[chat_id].excel_schedule_file_name}</b>\" успешно сохранен!',
                              reply_markup=markup_inline,
                              parse_mode='html')


# описание того, если пользователь откажется отправлять данные документа с расписанием
def deny_to_save_schedule_document(message):
    chat_id = message.chat.id

    markup_inline = types.InlineKeyboardMarkup()
    markup_inline.row(types.InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_schedule'),
                      types.InlineKeyboardButton(text='🔁 Повторить отправку',
                                                 callback_data='add_file_from_user_callback'))

    try:
        os.remove('./xlsx_files/' + str(users[chat_id].excel_schedule_file_name))

        bot.edit_message_text(chat_id=chat_id,
                              message_id=message.message_id,
                              text=f'Отмена записи файла \"<b>{users[chat_id].excel_schedule_file_name}</b>\".',
                              reply_markup=markup_inline,
                              parse_mode='html')

    except FileNotFoundError:
        files = glob.glob('./xlsx_files/*')
        for file in files:
            os.remove(file)
        bot.edit_message_text(chat_id=chat_id,
                              message_id=message.message_id,
                              text=f'Отмена записи файла \"<b>{users[chat_id].excel_schedule_file_name}</b>\".',
                              reply_markup=markup_inline,
                              parse_mode='html')


# запись данных из таблицы excel в базу данных
def schedule_document_to_db(message, past_day=None, past_month=None):
    chat_id = message.chat.id

    try:
        # подключение к базе данных
        with sqlite3.connect('./data_bases/schedule.sqlite3') as connection_to_db_main:
            cursor_db = connection_to_db_main.cursor()

        # удаление старых записей
        cursor_db.execute('''
            DELETE FROM 
                lessons_info 
            WHERE 
                (schedule_lesson_day < ('%s') AND schedule_lesson_month <= ('%s')) OR schedule_lesson_month < ('%s');
        '''
                          % (str(past_day), str(past_month), str(past_month))
                          )

        # подключение к таблице с расписанием
        main_table = openpyxl.load_workbook('./xlsx_files/' + str(users[chat_id].excel_schedule_file_name))
        worksheet = main_table.active

        max_column = worksheet.max_column

        # начало чтения информации
        start_row = 1
        while worksheet.cell(row=start_row, column=1).value != 'Дата':
            start_row += 1

        # конец чтения информации
        end_row = start_row
        while ((str(worksheet.cell(row=end_row, column=1).value) != 'None') or
               (str(worksheet.cell(row=end_row, column=2).value) != 'None') or
               (str(worksheet.cell(row=end_row, column=3).value) != 'None')):
            end_row += 3

        for current_platoon in users[chat_id].groups_in_xlsx_save.keys():
            for value in range(len(users[chat_id].groups_in_xlsx_save[str(current_platoon)])):
                current_group = users[chat_id].groups_in_xlsx_save[str(current_platoon)][value]

                cursor_db.execute('''
                    INSERT OR IGNORE INTO
                        group_info (
                            group_id,
                            schedule_course, 
                            schedule_platoon, 
                            schedule_group) 
                    VALUES (
                        (
                        SELECT 
                            group_id 
                        FROM 
                            group_info 
                        WHERE 
                            schedule_course == {course} AND schedule_platoon == '{platoon}' AND schedule_group == '{group}'
                        ), 
                        {course}, 
                        '{platoon}', 
                        '{group}');
                '''
                                  .format(course=int(last_num_year_of_start - int(current_platoon.split(' ')[0][-2]) + 1) % 10,
                                          platoon=str(current_platoon),
                                          group=str(current_group)))

                # проверка, есть ли записи в таблице групп базы данных schedule
                group_index = 1
                cursor_db.execute('''
                    SELECT 
                        group_id 
                    FROM 
                        group_info 
                    WHERE 
                        schedule_course == ('%d') AND schedule_platoon == ('%s') AND schedule_group == ('%s');
                '''
                                  % (int(last_num_year_of_start - int(current_platoon.split(' ')[0][-2]) + 1) % 10, str(current_platoon), str(current_group)))
                group_index = int(cursor_db.fetchone()[0])

                # list_time = []

                date_from_xlsx = ''
                time_from_xlsx = ''
                start_column_course = 3
                start_column_platoon = 3
                end_column_platoon = 4
                start_column_group = 3
                end_column_group = 4

                rasp_dictionary = {}

                # получение информации о начале чтения данных о занятиях
                for start_of_read in range(start_column_course, max_column + 1):
                    if str(worksheet.cell(row=start_row, column=start_of_read).value) == str(current_platoon):
                        start_column_platoon = start_of_read
                        end_column_platoon = start_of_read + 1
                        break
                while (str(worksheet.cell(row=start_row, column=end_column_platoon).value) == 'None' or
                       str(worksheet.cell(row=start_row + 1, column=end_column_platoon).value) == 'None'):
                    end_column_platoon += 1
                    if end_column_platoon == max_column + 1:
                        break
                end_column_platoon -= 1

                for start_of_read in range(start_column_platoon, max_column + 1):
                    if str(worksheet.cell(row=start_row + 1, column=start_of_read).value) == str(current_group):
                        start_column_group = start_of_read
                        end_column_group = start_of_read + 1
                        break
                while str(worksheet.cell(row=start_row + 1, column=end_column_group).value) == 'None':
                    end_column_group += 1
                    if end_column_group == max_column + 1:
                        break
                end_column_group -= 1

                # print(f'{start_column_course}, {max_column}, {start_column_platoon}, {end_column_platoon}, {start_column_group}, {end_column_group}')

                lesson_count_for_current_date = 1

                # непосредственно извлечение данных из таблицы и их запись в словарь "rasp_dictionary[date_from_xlsx][time_from_xlsx] = []"
                for current_string in range(start_row + 3, end_row, 3):
                    # if str(worksheet.cell(row=current_string, column=2).value) not in list_time:
                    #     if str(worksheet.cell(row=current_string, column=2).value) != 'None':
                    #         list_time.append(str(worksheet.cell(row=current_string, column=2).value))
                    #     else:
                    #         if 'None' in list_time:
                    #             list_time.append(f'{count}е занятие')
                    #         else:
                    #             list_time.append('None')
                    #             list_time.append(f'{count}е занятие')

                    if str(worksheet.cell(row=current_string, column=1).value) != 'None':  # есть дата
                        lesson_count_for_current_date = 1
                        date_from_xlsx = str(worksheet.cell(row=current_string, column=1).value).split(',')[0]
                        rasp_dictionary[date_from_xlsx] = {}
                        if str(worksheet.cell(row=current_string, column=2).value) != 'None':  # есть дата и есть время
                            time_from_xlsx = str(worksheet.cell(row=current_string, column=2).value)
                        else:  # есть дата и нет времени
                            time_from_xlsx = f'{lesson_count_for_current_date}е занятие'
                        rasp_dictionary[date_from_xlsx][time_from_xlsx] = []
                    else:
                        if str(worksheet.cell(row=current_string, column=2).value) != 'None':  # нет даты и есть время
                            time_from_xlsx = str(worksheet.cell(row=current_string, column=2).value)
                        else:  # нет даты и нет времени
                            time_from_xlsx = f'{lesson_count_for_current_date}е занятие'
                        rasp_dictionary[date_from_xlsx][time_from_xlsx] = []

                    is_group = True
                    is_platoon = True
                    is_course = True

                    # получение информации о том, является ли занятие общим для курса
                    for is_course_help in range(4, max_column + 1):
                        if str(worksheet.cell(row=current_string, column=is_course_help).value) != 'None':
                            is_course = False
                    if is_course:
                        if not (str(worksheet.cell(row=current_string, column=start_column_course).value) != 'None'
                                and str(worksheet.cell(row=current_string + 2, column=max_column).value) != 'None'):
                            is_course = False
                    # получение информации о том, является ли занятие общим для взвода
                    for is_platoon_help in range(start_column_platoon + 1, end_column_platoon + 1):
                        if str(worksheet.cell(row=current_string, column=is_platoon_help).value) != 'None':
                            is_platoon = False
                    if is_platoon:
                        if not (str(worksheet.cell(row=current_string, column=start_column_platoon).value) != 'None'
                                and str(worksheet.cell(row=current_string + 2, column=end_column_platoon).value) != 'None'):
                            is_platoon = False
                    # получение информации о том, является ли занятие общим для группы
                    for is_group_help in range(start_column_group + 1, end_column_group + 1):
                        if str(worksheet.cell(row=current_string, column=is_group_help).value) != 'None':
                            is_group = False
                    if is_group:
                        if not (str(worksheet.cell(row=current_string, column=start_column_group).value) != 'None'
                                and str(worksheet.cell(row=current_string + 2, column=end_column_group).value) != 'None'):
                            is_group = False

                    # занесение информации из .xlsx файла в словарь
                    if is_group:  # если занятие группы
                        rasp_dictionary[date_from_xlsx][time_from_xlsx].append(
                            str(worksheet.cell(row=current_string, column=start_column_group).value))
                        rasp_dictionary[date_from_xlsx][time_from_xlsx].append(
                            str(worksheet.cell(row=current_string + 1, column=start_column_group).value))
                        rasp_dictionary[date_from_xlsx][time_from_xlsx].append(
                            str(worksheet.cell(row=current_string + 2, column=start_column_group).value))
                        rasp_dictionary[date_from_xlsx][time_from_xlsx].append(
                            str(worksheet.cell(row=current_string + 2, column=end_column_group).value))
                    elif is_platoon:  # если занятие взвода
                        rasp_dictionary[date_from_xlsx][time_from_xlsx].append(
                            str(worksheet.cell(row=current_string, column=start_column_platoon).value))
                        rasp_dictionary[date_from_xlsx][time_from_xlsx].append(
                            str(worksheet.cell(row=current_string + 1, column=start_column_platoon).value))
                        rasp_dictionary[date_from_xlsx][time_from_xlsx].append(
                            str(worksheet.cell(row=current_string + 2, column=start_column_platoon).value))
                        rasp_dictionary[date_from_xlsx][time_from_xlsx].append(
                            str(worksheet.cell(row=current_string + 2, column=end_column_platoon).value))
                    elif is_course:  # если занятие курса
                        rasp_dictionary[date_from_xlsx][time_from_xlsx].append(
                            str(worksheet.cell(row=current_string, column=start_column_course).value))
                        rasp_dictionary[date_from_xlsx][time_from_xlsx].append(
                            str(worksheet.cell(row=current_string + 1, column=start_column_course).value))
                        rasp_dictionary[date_from_xlsx][time_from_xlsx].append(
                            str(worksheet.cell(row=current_string + 2, column=start_column_course).value))
                        rasp_dictionary[date_from_xlsx][time_from_xlsx].append(
                            str(worksheet.cell(row=current_string + 2, column=max_column).value))
                    else:
                        del rasp_dictionary[date_from_xlsx][time_from_xlsx]

                    lesson_count_for_current_date += 1
                # запись данных расписания в таблицу занятий базы данных schedule
                for current_date in rasp_dictionary.keys():
                    print((str(current_date)).split('.')[0], past_day, (str(current_date)).split('.')[1], past_month)
                    if (((str(current_date)).split('.')[0] >= past_day and (str(current_date)).split('.')[1] >= past_month)
                            or (str(current_date)).split('.')[1] > past_month):

                        cursor_db.execute('''
                            DELETE FROM
                                lessons_info
                            WHERE
                                group_id == ('%d') AND schedule_lesson_day == ('%s') AND schedule_lesson_month == ('%s');
                        '''
                                          % (int(group_index), str(current_date).split('.')[0], str(current_date).split('.')[1]))

                        for time_for_current_date in rasp_dictionary[current_date].keys():
                            cursor_db.execute('''
                                INSERT INTO 
                                    lessons_info (
                                        group_id, 
                                        schedule_lesson_day, 
                                        schedule_lesson_month, 
                                        schedule_lesson_time, 
                                        schedule_lesson, 
                                        schedule_type_of_lesson, 
                                        schedule_tutor, 
                                        schedule_lesson_room) 
                                VALUES ('%d', '%s', '%s', '%s', '%s', '%s', '%s', '%s');
                            '''
                                              % (int(group_index), str(current_date).split('.')[0],
                                                 str(current_date).split('.')[1],
                                                 str(time_for_current_date),
                                                 str(rasp_dictionary[current_date][time_for_current_date][0]),
                                                 str(rasp_dictionary[current_date][time_for_current_date][1]),
                                                 str(rasp_dictionary[current_date][time_for_current_date][2]),
                                                 str(rasp_dictionary[current_date][time_for_current_date][3])))

        # проверка наличия удаленных записей занятий в соответствующих группах
        cursor_db.execute('''
            DELETE FROM
                group_info
            WHERE
                group_id IN(
                    SELECT
                        gi.group_id
                    FROM
                        group_info AS gi
                    LEFT OUTER JOIN
                        lessons_info AS li
                    ON
                        gi.group_id = li.group_id
                    WHERE
                        lesson_id IS NULL
                    GROUP BY 
                        gi.group_id);
        ''')

        connection_to_db_main.commit()
        cursor_db.close()
        connection_to_db_main.close()

    except Exception as e:
        return error_from_user(message=message, error_code=e)


def create_schedule_courses(message, do_edit_message=False, process_index=1):
    chat_id = message.chat.id

    try:
        with sqlite3.connect('./data_bases/schedule.sqlite3') as connection_to_db_main:
            cursor_db = connection_to_db_main.cursor()
        cursor_db.execute('''
            SELECT 
                schedule_course 
            FROM 
                group_info 
            GROUP BY 
                schedule_course 
            ORDER BY 
                schedule_course;
        ''')
        list_kurs_from_schedule = cursor_db.fetchall()
        connection_to_db_main.commit()
        cursor_db.close()
        connection_to_db_main.close()

        button_course = []
        help_list = []
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                           one_time_keyboard=True,
                                           row_width=5,
                                           input_field_placeholder='Выберите нужный курс...')
        for current_button in list_kurs_from_schedule:
            button_course.append(types.KeyboardButton(text=str(current_button[0])))
            help_list.append(str(current_button[0]))
        markup.add(*button_course)
        markup.add(types.KeyboardButton(text='🔙 Назад'))
        markup.add(types.KeyboardButton(text='🏠 На главную'))

        if do_edit_message:
            bot.delete_message(chat_id=chat_id,
                               message_id=bot.edit_message_text(chat_id=chat_id,
                                                                message_id=message.message_id,
                                                                text='⏳ Подготовка запроса...',
                                                                parse_mode='html').message_id)
            bot.send_message(chat_id=chat_id,
                             text=f'{process_index_to_emogi[1][process_index - 1]} У меня есть информация по слeдующим <u><b>КУРСАМ</b></u>: ',
                             parse_mode='html',
                             reply_markup=markup)

        else:
            bot.send_message(chat_id=chat_id,
                             text=f'{process_index_to_emogi[1][process_index - 1]} У меня есть информация по слeдующим <u><b>КУРСАМ</b></u>: ',
                             parse_mode='html',
                             reply_markup=markup)

        bot.register_next_step_handler(message, check_schedule_course, list_kurs=help_list,
                                       process_index=process_index)

    except Exception as e:
        return error_from_user(message=message, error_code=e)


def check_schedule_course(message, help_course=None, list_kurs=None, func_input_type=None, process_index=1):
    chat_id = message.chat.id

    # подключение к базе данных
    if (message.text in list_kurs) or (func_input_type == 'back'):
        input_kurs = '1'
        if func_input_type == 'back':
            input_kurs = help_course
        elif message.text in list_kurs:
            input_kurs = message.text
        with sqlite3.connect('./data_bases/schedule.sqlite3') as connection_to_db_main:
            cursor_db = connection_to_db_main.cursor()
        cursor_db.execute('''
            SELECT 
                schedule_platoon 
            FROM 
                group_info 
            WHERE 
                schedule_course == ('%d')
            GROUP BY 
                schedule_platoon 
            ORDER BY 
                schedule_platoon;
        '''
                          % (int(input_kurs)))
        list_vzvod_from_schedule = cursor_db.fetchall()
        connection_to_db_main.commit()
        cursor_db.close()
        connection_to_db_main.close()

        button_platoon = []
        help_list = []
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                           one_time_keyboard=True,
                                           row_width=5,
                                           input_field_placeholder='Выберите нужный взвод...')
        for current_button in list_vzvod_from_schedule:
            button_platoon.append(types.KeyboardButton(text=str(current_button[0])))
            help_list.append(str(current_button[0]))
        markup.add(*button_platoon)
        markup.add(types.KeyboardButton(text='🔙 Назад'))
        markup.add(types.KeyboardButton(text='🏠 На главную'))

        bot.reply_to(message=message,
                     text=f'{process_index_to_emogi[1][process_index - 1]} У меня есть информация по слeдующим <u><b>ВЗВОДАМ</b></u>: ',
                     reply_markup=markup,
                     parse_mode='html')

        bot.register_next_step_handler(message, check_schedule_platoon, help_course=input_kurs, list_platoon=help_list,
                                       process_index=process_index)
    elif message.text == '🔙 Назад':
        delete_reply_markup(message, text_hint='Возвращаю на страницу расписания...')
        return command_schedule(message, do_edit_message=False)
    elif message.text == '🏠 На главную':
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message)
    elif message.text in command_list:
        return all_commands(message=message)
    else:
        bot.send_message(chat_id=chat_id,
                         text=f'ℹ Курса с таким номером не найдено, введите, пожалуйста, корректные данные или обратитесь за обратной связью: ',
                         reply_markup=links_without_button_back())
        bot.register_next_step_handler(message, check_schedule_course, list_kurs=list_kurs,
                                       process_index=process_index)


def check_schedule_platoon(message, help_course=None, help_platoon=None, list_platoon=None, func_input_type=None,
                                process_index=1):
    chat_id = message.chat.id

    # подключение к базе данных
    if (message.text in list_platoon) or (func_input_type == 'back'):
        input_vzvod = '931'
        if func_input_type == 'back':
            input_vzvod = help_platoon
        elif message.text in list_platoon:
            input_vzvod = message.text
        with sqlite3.connect('./data_bases/schedule.sqlite3') as connection_to_db_main:
            cursor_db = connection_to_db_main.cursor()
        cursor_db.execute('''
            SELECT 
                schedule_group 
            FROM 
                group_info 
            WHERE 
                schedule_platoon == ('%s')
            GROUP BY 
                schedule_group 
            ORDER BY
                schedule_group;
        '''
                          % (str(input_vzvod)))
        list_group_from_schedule = cursor_db.fetchall()
        connection_to_db_main.commit()
        cursor_db.close()
        connection_to_db_main.close()

        button_group = []
        help_list = []
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                           one_time_keyboard=True,
                                           row_width=5,
                                           input_field_placeholder='Выберите нужную группу...')
        for current_button in list_group_from_schedule:
            button_group.append(types.KeyboardButton(text=str(current_button[0])))
            help_list.append(str(current_button[0]))
        markup.add(*button_group)
        markup.add(types.KeyboardButton(text='🔙 Назад'))
        markup.add(types.KeyboardButton(text='🏠 На главную'))

        bot.reply_to(message=message,
                     text=f'{process_index_to_emogi[1][process_index - 1]} У меня есть информация по слeдующим <u><b>ГРУППАМ</b></u>: ',
                     reply_markup=markup,
                     parse_mode='html')

        bot.register_next_step_handler(message, check_schedule_group, help_course=help_course, help_platoon=input_vzvod,
                                       list_group=help_list, process_index=process_index)
    elif message.text == '🔙 Назад':
        return create_schedule_courses(message, do_edit_message=False, process_index=process_index)
    elif message.text == '🏠 На главную':
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message)
    elif message.text in command_list:
        return all_commands(message=message)
    else:
        bot.send_message(chat_id=chat_id,
                         text=f'ℹ Взвода с таким номером не найдено, введите, пожалуйста, корректные данные или обратитесь за обратной связью: ',
                         reply_markup=links_without_button_back())
        bot.register_next_step_handler(message, check_schedule_platoon, help_course=help_course, list_platoon=list_platoon,
                                       process_index=process_index)


# вывод информации о расписании для группы определенного взвода на сегодня и завтра
def check_schedule_group(message, help_course=None, help_platoon=None, help_group=None, list_group=None,
                              process_index=1):
    chat_id = message.chat.id

    # подключение к базе данных
    if (message.text in list_group) or (help_group):
        if process_index == 1:
            now_day = str(datetime.now().day)
            if len(now_day) == 1:
                now_day = '0' + now_day
            now_month = str(datetime.now().month)
            if len(now_month) == 1:
                now_month = '0' + now_month
            tomorrow_day = str((datetime.now() + timedelta(days=1)).day)
            if len(tomorrow_day) == 1:
                tomorrow_day = '0' + tomorrow_day
            tomorrow_month = str((datetime.now() + timedelta(days=1)).month)
            if len(tomorrow_month) == 1:
                tomorrow_month = '0' + tomorrow_month

            input_group = message.text
            with sqlite3.connect('./data_bases/schedule.sqlite3') as connection_to_db_main:
                cursor_db = connection_to_db_main.cursor()

            cursor_db.execute('''
                SELECT 
                    li.schedule_lesson_time, 
                    li.schedule_lesson, 
                    li.schedule_type_of_lesson, 
                    li.schedule_tutor, 
                    li.schedule_lesson_room
                FROM 
                    lessons_info as li
                LEFT OUTER JOIN 
                    group_info AS gi
                ON 
                    li.group_id = gi.group_id
                WHERE 
                    li.schedule_lesson_day == ('%s') AND li.schedule_lesson_month == ('%s') AND schedule_course == ('%d') AND schedule_platoon == ('%s') AND schedule_group == ('%s')
                ORDER BY 
                    li.schedule_lesson_month, li.schedule_lesson_time;
            '''
                              % (str(now_day), str(now_month), int(help_course), str(help_platoon), str(input_group)))
            lessons_today = cursor_db.fetchall()

            cursor_db.execute('''
                SELECT 
                    li.schedule_lesson_time, 
                    li.schedule_lesson, 
                    li.schedule_type_of_lesson, 
                    li.schedule_tutor, 
                    li.schedule_lesson_room
                FROM 
                    lessons_info as li
                LEFT OUTER JOIN 
                    group_info AS gi
                ON 
                    li.group_id = gi.group_id
                WHERE 
                    li.schedule_lesson_day == ('%s') AND li.schedule_lesson_month == ('%s') AND schedule_course == ('%d') AND schedule_platoon == ('%s') AND schedule_group == ('%s')
                ORDER BY 
                    li.schedule_lesson_month, 
                    li.schedule_lesson_time;
            '''
                              % (str(tomorrow_day), str(tomorrow_month), int(help_course), str(help_platoon), str(input_group)))
            lessons_tomorrow = cursor_db.fetchall()

            connection_to_db_main.commit()
            cursor_db.close()
            connection_to_db_main.close()

            lessons_count = 1
            info_lessons = (
                f'{days_week_to_emoji[datetime.now().weekday()]} <i><b><u>Расписание на сегодня</u></b></i>, {num_to_emoji(str(now_day))}.{num_to_emoji(str(now_month))}, {days_week[datetime.now().weekday()]}, '
                f'<i><b><u>для</u></b></i> {num_to_emoji(str(help_course))}<i><b><u>го курса</u></b></i> {num_to_emoji(str(help_platoon).split(" ")[0])}-{num_to_emoji(str(input_group).split("/")[-1])}<i><b><u>го взвода:</u></b></i>\n\n')
            if lessons_today:
                for current_lesson in lessons_today:
                    info_lessons += (f'{num_to_emoji(str(lessons_count))} 🕰 \t{current_lesson[0]} 🕰\n'
                                     f'🎓 \t<b>{current_lesson[1]}</b>\n'
                                     f'🔬 \t{current_lesson[2]} / 👨 {current_lesson[3]}\n'
                                     f'🚪 \t{current_lesson[4]}\n\n')
                    lessons_count += 1
                lessons_count = 1
            else:
                info_lessons += '🧧 Нет информации о занятиях!\n\n'

            info_lessons += '\n'

            info_lessons += (
                f'{days_week_to_emoji[(datetime.now() + timedelta(days=1)).weekday()]} <i><b><u>Расписание на завтра</u></b></i>, {num_to_emoji(str(tomorrow_day))}.{num_to_emoji(str(tomorrow_month))}, {days_week[(datetime.now() + timedelta(days=1)).weekday()]}, '
                f'<i><b><u>для</u></b></i> {num_to_emoji(str(help_course))}<i><b><u>го курса</u></b></i> {num_to_emoji(str(help_platoon).split(" ")[0])}-{num_to_emoji(str(input_group).split("/")[-1])}<i><b><u>го взвода:</u></b></i>\n\n')
            if lessons_tomorrow:
                for current_lesson in lessons_tomorrow:
                    info_lessons += (f'{num_to_emoji(str(lessons_count))} 🕰 \t{current_lesson[0]} 🕰\n'
                                     f'🎓 \t<b>{current_lesson[1]}</b>\n'
                                     f'🔬 \t{current_lesson[2]} / 👨 {current_lesson[3]}\n'
                                     f'🚪 \t{current_lesson[4]}\n\n')
                    lessons_count += 1
            else:
                info_lessons += '🧧 Нет информации о занятиях!\n\n'

            bot.send_message(chat_id=chat_id,
                             text=info_lessons,
                             parse_mode='html')
            return check_schedule_platoon(message=message, help_course=help_course, help_platoon=help_platoon,
                                               list_platoon=[], func_input_type='back', process_index=process_index)
        elif process_index == 2:
            if len(schedule.get_jobs()) == 0:
                schedule.every().day.at(send_time).do(subscribe_to_schedule, help_course=help_course, help_platoon=help_platoon,
                                                      help_group=message.text, chat_id=chat_id)
                Thread(target=schedule_checker).start()
                user = users[chat_id]
                user.is_subscribe_to_schedule = True
                delete_reply_markup(message=message)
                bot.send_message(chat_id=chat_id,
                                 text=f'Вы подписались на ежедневные уведомления о занятиях для {help_course}го курса {help_platoon}/{message.text}го взвода.')
                return command_schedule(message=message, do_edit_message=False)

    elif message.text == '🔙 Назад':
        return check_schedule_course(message, help_course=help_course, list_kurs=[], func_input_type='back',
                                          process_index=process_index)
    elif message.text == '🏠 На главную':
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message)
    elif message.text in command_list:
        return all_commands(message=message)
    else:
        bot.send_message(chat_id=chat_id,
                         text=f'ℹ Группы с таким номером не найдено, введите, пожалуйста, корректные данные или обратитесь за обратной связью: ',
                         reply_markup=links_without_button_back())
        bot.register_next_step_handler(message, check_schedule_group, help_course=help_course, help_platoon=help_platoon,
                                       list_group=list_group, process_index=process_index)


# раздел "получение информации о пользователе"
def command_all(message, do_edit_message=False, process_index=1):
    chat_id = message.chat.id

    with sqlite3.connect('./data_bases/students.sqlite3') as connection_to_db_main:
        cursor_db = connection_to_db_main.cursor()
    cursor_db.execute('''
        SELECT 
            course 
        FROM 
            cadets 
        GROUP BY 
            course 
        ORDER BY 
            course;
    ''')
    list_kurs_from_users = cursor_db.fetchall()
    connection_to_db_main.commit()
    cursor_db.close()
    connection_to_db_main.close()

    button_course = []
    help_list = []
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                       one_time_keyboard=True,
                                       row_width=5,
                                       input_field_placeholder='Выберите нужный курс...')
    for current_button in list_kurs_from_users:
        button_course.append(types.KeyboardButton(text=str(current_button[0])))
        help_list.append(str(current_button[0]))
    markup.add(*button_course)
    if process_index == 1:
        markup.add(types.KeyboardButton(text='🍉 За факультет'))
    if (process_index == 1) or (process_index == 2):
        markup.add(types.KeyboardButton(text='🏠 На главную'))
    else:
        markup.add(types.KeyboardButton(text='🔙 Назад'))

    if do_edit_message:
        bot.delete_message(chat_id=chat_id,
                           message_id=bot.edit_message_text(chat_id=chat_id,
                                                            message_id=message.message_id,
                                                            text='⏳ Подготовка запроса...',
                                                            parse_mode='html').message_id)
        bot.send_message(chat_id=chat_id,
                         text=f'{process_index_to_emogi[0][process_index - 1]} У меня есть информация по слeдующим <u><b>КУРСАМ</b></u>: ',
                         parse_mode='html',
                         reply_markup=markup)

    else:
        bot.reply_to(message=message,
                     text=f'{process_index_to_emogi[0][process_index - 1]} У меня есть информация по слeдующим <u><b>КУРСАМ</b></u>: ',
                     parse_mode='html',
                     reply_markup=markup)

    bot.register_next_step_handler(message, check_course, list_kurs=help_list,
                                   process_index=process_index)


# выбор курса в разделе "получение информации о пользователе"
def check_course(message, help_course=None, list_kurs=None, func_input_type=None, process_index=1):
    chat_id = message.chat.id

    # подключение к базе данных
    if (message.text in list_kurs) or (func_input_type == 'back'):
        if func_input_type != 'back':
            help_course = message.text
        with sqlite3.connect('./data_bases/students.sqlite3') as connection_to_db_main:
            cursor_db = connection_to_db_main.cursor()
        cursor_db.execute('''
            SELECT 
                platoon 
            FROM 
                cadets 
            WHERE 
                course == ('%d') 
            GROUP BY 
                platoon 
            ORDER BY 
                platoon;
        '''
                          % (int(help_course)))
        list_platoon_from_users = cursor_db.fetchall()
        connection_to_db_main.commit()
        cursor_db.close()
        connection_to_db_main.close()

        button_platoon = []
        help_list = []
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                           one_time_keyboard=True,
                                           row_width=5,
                                           input_field_placeholder='Выберите нужный взвод...')
        for current_button in list_platoon_from_users:
            button_platoon.append(types.KeyboardButton(text=str(current_button[0])))
            help_list.append(str(current_button[0]))
        markup.add(*button_platoon)
        if process_index == 1:
            markup.add(types.KeyboardButton(text='🍍 За курс'))
        elif process_index == 3:
            markup.add(types.KeyboardButton(text='🍍 Удалить курс'))
        markup.add(types.KeyboardButton(text='🔙 Назад'), types.KeyboardButton(text='🏠 На главную'))

        bot.reply_to(message=message,
                     text=f'{process_index_to_emogi[0][process_index - 1]} У меня есть информация по слeдующим <u><b>ВЗВОДАМ</b></u>: ',
                     reply_markup=markup,
                     parse_mode='html')

        bot.register_next_step_handler(message, check_platoon, help_course=help_course,
                                       list_platoon=help_list, process_index=process_index)
    elif message.text == '🍉 За факультет' and process_index == 1:
        return create_all_buttons(message=message, choice_info='faculty')
    elif (message.text == '🔙 Назад') and ((process_index == 3) or (process_index == 4)):
        delete_reply_markup(message=message, text_hint='Возвращаю на страницу изменения информации о курсантах...')
        return update_students(message=message, do_edit_message=False)
    elif message.text == '🏠 На главную':
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message)
    elif message.text in command_list:
        return all_commands(message=message)
    else:
        bot.send_message(chat_id=chat_id,
                         text=f'ℹ Курса с таким номером не найдено, введите, пожалуйста, корректные данные или обратитесь за обратной связью: ',
                         reply_markup=links_without_button_back())
        bot.register_next_step_handler(message, check_course, list_kurs=list_kurs,
                                       process_index=process_index)


# выбор взвода в разделе "получение информации о пользователе"
def check_platoon(message, help_course=None, help_platoon=None, list_platoon=None, func_input_type=None,
                                     process_index=1):
    chat_id = message.chat.id

    # подключение к базе данных
    if (message.text in list_platoon) or (func_input_type == 'back'):
        if func_input_type != 'back':
            help_platoon = message.text
        if process_index == 1:
            return create_all_buttons(message=message, help_course=help_course, help_platoon=help_platoon, choice_info='platoon')
        elif (process_index == 2) or (process_index == 3) or (process_index == 4):
            # подключение к базе данных
            with sqlite3.connect('./data_bases/students.sqlite3') as connection_to_db_main:
                cursor_db = connection_to_db_main.cursor()
            cursor_db.execute('''
                SELECT 
                    cadet 
                FROM 
                    cadets 
                WHERE 
                    course == ('%d') AND platoon == ('%d') 
                GROUP BY 
                    cadet 
                ORDER BY 
                    cadet;
            '''
                              % (int(help_course), int(help_platoon)))
            list_man_from_users = cursor_db.fetchall()
            connection_to_db_main.commit()
            cursor_db.close()
            connection_to_db_main.close()

            button_man = []
            help_list = []
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                               one_time_keyboard=True,
                                               row_width=3,
                                               input_field_placeholder='Выберите нужного курсанта...')
            for current_button in list_man_from_users:
                button_man.append(types.KeyboardButton(text=str(current_button[0])))
                help_list.append(str(current_button[0]))
            markup.add(*button_man)
            if process_index == 3:
                markup.add(types.KeyboardButton(text='🍈 Удалить взвод'))
            markup.add(types.KeyboardButton(text='🔙 Назад'), types.KeyboardButton(text='🏠 На главную'))

            bot.reply_to(message=message,
                         text=f'{process_index_to_emogi[0][process_index - 1]} У меня есть информация по слeдующим <u><b>КУРСАНТАМ</b></u>: ',
                         reply_markup=markup,
                         parse_mode='html')

            bot.register_next_step_handler(message, check_cadet, help_course=help_course, help_platoon=help_platoon,
                                           list_man=help_list, process_index=process_index)
    elif message.text == '🍍 За курс' and process_index == 1:
        return create_all_buttons(message=message, help_course=help_course, choice_info='course')
    elif message.text == '🍍 Удалить курс' and process_index == 3:
        return delete_cadet_first(message=message, help_course=help_course)
    elif message.text == '🔙 Назад':
        return command_all(message, do_edit_message=False, process_index=process_index)
    elif message.text == '🏠 На главную':
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message)
    elif message.text in command_list:
        return all_commands(message=message)
    else:
        bot.send_message(chat_id=chat_id,
                         text=f'ℹ Взвода с таким номером не найдено, введите, пожалуйста, корректные данные или обратитесь за обратной связью: ',
                         reply_markup=links_without_button_back())
        bot.register_next_step_handler(message, check_platoon, help_course=help_course,
                                       list_platoon=list_platoon, process_index=process_index)


def check_cadet(message, help_course=None, help_platoon=None, help_cadet=None, list_man=None, func_input_type=None, process_index=1):
    chat_id = message.chat.id

    if (message.text in list_man) or (func_input_type == 'back'):
        if func_input_type != 'back':
            help_cadet = message.text
        if process_index == 2:
            return create_add_cadet_buttons(message=message, help_course=help_course, help_platoon=help_platoon,
                                            help_cadet=help_cadet)
        elif process_index == 3:
            return delete_cadet_first(message=message, help_course=help_course, help_platoon=help_platoon,
                                      help_cadet=help_cadet)
        elif process_index == 4:
            return command_update_cadet(message=message, help_course=help_course, help_platoon=help_platoon,
                                      help_cadet=help_cadet)
    elif (message.text == '🍈 Удалить взвод') and (process_index == 3):
        return delete_cadet_first(message=message, help_course=help_course, help_platoon=help_platoon)
    elif message.text == '🔙 Назад':
        return check_course(message, help_course=help_course, list_kurs=[], func_input_type='back',
                                            process_index=process_index)
    elif message.text == '🏠 На главную':
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message)
    elif message.text in command_list:
        return all_commands(message=message)
    else:
        bot.send_message(chat_id=chat_id,
                         text=f'ℹ Курсанта с таким именем не найдено, введите, пожалуйста, корректные данные или обратитесь за обратной связью: ',
                         reply_markup=links_without_button_back())
        bot.register_next_step_handler(message, check_cadet, help_course=help_course, help_platoon=help_platoon,
                                       list_man=list_man, process_index=process_index)


def command_update_cadet(message, help_course=None, help_platoon=None, help_cadet=None):
    chat_id = message.chat.id
    users[chat_id].data_set.clear()

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                       one_time_keyboard=True,
                                       row_width=2,
                                       input_field_placeholder='Введите обновленные данные...')
    markup.add(types.KeyboardButton(text='🔙 Назад'), types.KeyboardButton(text='🏠 На главную'))

    bot.reply_to(message=message,
                 text=f'👤 Выбран курсант <u><i>{help_course}</i></u>го курса <u><i>{help_platoon}</i></u>го взвода <u><i>{help_cadet}</i></u>.\n\n'
                      f''
                      f'{update_student_rules}',
                 parse_mode='html',
                 reply_markup=markup)

    bot.register_next_step_handler(message, update_cadet_check, help_course=help_course, help_platoon=help_platoon, help_cadet=help_cadet)


def update_cadet_check(message, help_course=None, help_platoon=None, help_cadet=None):
    chat_id = message.chat.id

    if message.text in command_list:
        return all_commands(message=message)
    elif message.text == '🔙 Назад':
        return check_platoon(message, help_course=help_course, help_platoon=help_platoon, list_platoon=[], func_input_type='back', process_index=4)
    elif message.text == '🏠 На главную':
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message)
    elif message.text:
        info = ''

        " ".join(message.text.split(' ')).strip()
        info_about_student = [x for x in message.text.split(' ') if x]
        if len(info_about_student) == 6:
            error_info = ''
            help_list = []

            if (len(info_about_student[0].lstrip('0')) == 1) and info_about_student[0].isdigit():
                help_list.append(int(info_about_student[0].lstrip('0')))  # int
            else:
                error_info += error_add_student[0]

            if ((len(info_about_student[1].lstrip('0')) >= 3) and (len(info_about_student[1].lstrip('0')) <= 4)
                    and (info_about_student[1][-1] != '0') and info_about_student[1].isdigit()):
                help_list.append(int(info_about_student[1].lstrip('0')))  # int
            else:
                error_info += error_add_student[1]

            if (len(info_about_student[2]) >= 2) and (len(info_about_student[3]) >= 2) and (
                    info_about_student[2][0] not in denied_symbols) and (
                    info_about_student[3][0] not in denied_symbols):
                help_list.append(info_about_student[2] + ' ' + info_about_student[3])  # str
            else:
                error_info += error_add_student[2]

            if info_about_student[4] == '_':
                help_list.append(default_cadet_value)  # int
            elif info_about_student[4].isdigit():
                if int(info_about_student[4]) in range(1, 3000):
                    help_list.append(int(info_about_student[4]))  # int
            else:
                error_info += error_add_student[3]

            if info_about_student[5] == '_':
                help_list.append(default_cadet_value)  # int
            elif info_about_student[5].isdigit():
                if int(info_about_student[5]) in range(1, 3000):
                    help_list.append(int(info_about_student[5]))  # int
            else:
                error_info += error_add_student[4]

            if len(help_list) == 5:
                with sqlite3.connect('./data_bases/students.sqlite3') as connection_to_db:
                    cursor_db = connection_to_db.cursor()

                cursor_db.execute('''
                    SELECT 
                        c.course,
                        c.platoon,
                        c.cadet,
                        p.grade,
                        p.discipline 
                    FROM 
                        cadets AS c
                    LEFT OUTER JOIN
                        parameters AS p
                    ON
                        c.student_id = p.student_id
                    WHERE 
                        course == ('%d') AND platoon == ('%d') AND cadet == ('%s');
                '''
                                  % (int(help_course), int(help_platoon), str(help_cadet)))
                user_from_db = cursor_db.fetchone()
                connection_to_db.commit()
                cursor_db.close()
                connection_to_db.close()

                users[chat_id].data_set = help_list
                info += (f'Курс\n'
                         f'\t└ <u><b>{user_from_db[0]}</b></u>й ➡ <u><b>{help_list[0]}</b></u>й\n'
                         f'Взвод\n'
                         f'\t└ <u><b>{user_from_db[1]}</b></u>й ➡ <u><b>{help_list[1]}</b></u>й\n'
                         f'Курсант\n'
                         f'\t└ <u><b>{user_from_db[2]}</b></u> ➡ <u><b>{help_list[2]}</b></u>\n'
                         f'Успеваемость\n'
                         f'\t└ <u><b>{user_from_db[3]}</b></u> ➡ <u><b>{help_list[3]}</b></u>\n'
                         f'Дисциплина\n'
                         f'\t└ <u><b>{user_from_db[4]}</b></u> ➡ <u><b>{help_list[4]}</b></u>\n')
            else:
                info += (
                    f'Неверно введены следующие поля (<u>{info_about_student[2]} {info_about_student[3]} ({info_about_student[0]}, {info_about_student[1]}</u>)):\n'
                    f'<b>{error_info}</b>\n')
        elif len(info_about_student) < 6:
            info += 'Заполнены не все поля!\n\n'
        else:
            info += 'Указана избыточная информация!\n\n'

        if users[chat_id].data_set:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                               one_time_keyboard=True,
                                               row_width=2,
                                               input_field_placeholder='Повторить?')
            markup.add(types.KeyboardButton(text='✔ Обновить'), types.KeyboardButton(text='❌ Не обновлять'))

            bot.send_message(chat_id=chat_id,
                             text=f'Вы хотите обновить данные в следующем виде:\n\n'
                                  f''
                                  f'{info}'
                                  f''
                                  f'Обновить данные?',
                             parse_mode='html',
                             reply_markup=markup)
        else:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                               one_time_keyboard=True,
                                               row_width=2,
                                               input_field_placeholder='Повторить?')
            markup.add(types.KeyboardButton(text='🔁 Повторить'), types.KeyboardButton(text='🏠 На главную'))

            bot.send_message(chat_id=chat_id,
                             text=f'{info}\n'
                                  f''
                                  f'Повторить отправку?',
                             parse_mode='html',
                             reply_markup=markup)

        bot.register_next_step_handler(message, do_update_cadet, help_course=help_course, help_platoon=help_platoon, help_cadet=help_cadet)

    else:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                           one_time_keyboard=True,
                                           row_width=2,
                                           input_field_placeholder='Повторить?')
        markup.add(types.KeyboardButton(text='🔙 Назад'), types.KeyboardButton(text='🏠 На главную'))

        bot.send_message(chat_id=chat_id,
                         text='ℹ Невозможно определить действие. Введите, пожалуйста, корректные данные или используйте навигационные кнопки ниже.',
                         parse_mode='html',
                         reply_markup=markup)

        bot.register_next_step_handler(message, update_cadet_check, help_course=help_course, help_platoon=help_platoon,
                                       help_cadet=help_cadet)


def do_update_cadet(message, help_course=None, help_platoon=None, help_cadet=None):
    chat_id = message.chat.id

    if message.text in command_list:
        users[chat_id].data_set.clear()
        return all_commands(message=message)
    elif message.text == '🔙 Назад':
        users[chat_id].data_set.clear()
        return check_platoon(message, help_course=help_course, help_platoon=help_platoon, list_platoon=[], func_input_type='back', process_index=4)
    elif message.text == '🏠 На главную':
        users[chat_id].data_set.clear()
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message)
    elif message.text == '✔ Обновить':
        with sqlite3.connect('./data_bases/students.sqlite3') as connection_to_db:
            cursor_db = connection_to_db.cursor()
        cursor_db.execute('''
            UPDATE
                parameters
            SET
                grade = ('%d'),
                discipline = ('%d')
            WHERE
                student_id == (
                    SELECT
                        student_id
                    FROM
                        cadets
                    WHERE
                        course == ('%d') AND platoon == ('%d') AND cadet == ('%s')
                );
        '''
                          % (int(users[chat_id].data_set[3]), int(users[chat_id].data_set[4]),
                             int(help_course), int(help_platoon), str(help_cadet)))
        cursor_db.execute('''
            UPDATE
                cadets
            SET
                course = ('%d'),
                platoon = ('%d'),
                cadet = ('%s')
            WHERE 
                course == ('%d') AND platoon == ('%d') AND cadet == ('%s');
        '''
                          % (int(users[chat_id].data_set[0]), int(users[chat_id].data_set[1]), str(users[chat_id].data_set[2]),
                             int(help_course), int(help_platoon), str(help_cadet)))
        connection_to_db.commit()
        cursor_db.close()
        connection_to_db.close()

        users[chat_id].data_set.clear()

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                           one_time_keyboard=True,
                                           row_width=2,
                                           input_field_placeholder='Обновить?')
        markup.add(types.KeyboardButton(text='🔁 Обновить'), types.KeyboardButton(text='🏠 На главную'))
        bot.send_message(chat_id=chat_id,
                         text='✔ Информация обновлена. Хотите обновить информацию о других курсантах?',
                         parse_mode='html',
                         reply_markup=markup)
        bot.register_next_step_handler(message, do_update_cadet_after_input, help_course=help_course,
                                       help_platoon=help_platoon, help_cadet=help_cadet)
    elif message.text == '❌ Не обновлять':
        users[chat_id].data_set.clear()

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                           one_time_keyboard=True,
                                           row_width=2,
                                           input_field_placeholder='Обновить?')
        markup.add(types.KeyboardButton(text='🔁 Обновить'), types.KeyboardButton(text='🏠 На главную'))
        bot.send_message(chat_id=chat_id,
                         text='❌ Отмена обновления информации. Хотите обновить информацию о других курсантах?',
                         parse_mode='html',
                         reply_markup=markup)
        bot.register_next_step_handler(message, do_update_cadet_after_input, help_course=help_course,
                                       help_platoon=help_platoon, help_cadet=help_cadet)

    elif message.text == '🔁 Повторить':
        users[chat_id].data_set.clear()
        return command_update_cadet(message=message, help_course=help_course, help_platoon=help_platoon,
                                    help_cadet=help_cadet)
    else:
        bot.send_message(chat_id=chat_id,
                         text='ℹ Невозможно определить действие. Введите, пожалуйста, корректные данные или используйте навигационные кнопки ниже.',
                         parse_mode='html')

        bot.register_next_step_handler(message, do_update_cadet, help_course=help_course, help_platoon=help_platoon, help_cadet=help_cadet)


def do_update_cadet_after_input(message, help_course=None, help_platoon=None, help_cadet=None):
    chat_id = message.chat.id

    if message.text in command_list:
        return all_commands(message=message)
    elif message.text == '🔁 Обновить':
        return check_platoon(message, help_course=help_course, help_platoon=help_platoon, list_platoon=[],
                             func_input_type='back', process_index=4)
    elif message.text == '🏠 На главную':
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message)
    else:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                           one_time_keyboard=True,
                                           row_width=2,
                                           input_field_placeholder='Обновить?')
        markup.add(types.KeyboardButton(text='🔁 Обновить'), types.KeyboardButton(text='🏠 На главную'))

        bot.send_message(chat_id=chat_id,
                         text='ℹ Невозможно определить действие. Введите, пожалуйста, корректные данные или используйте навигационные кнопки ниже.',
                         parse_mode='html',
                         reply_markup=markup)

        bot.register_next_step_handler(message, do_update_cadet_after_input, help_course=help_course, help_platoon=help_platoon,
                                           help_cadet=help_cadet)


def delete_cadet_first(message, help_course=None, help_platoon=None, help_cadet=None):
    chat_id = message.chat.id

    info = ''
    if help_cadet:
        info = (f'❓ Вы действительно хотите удалить информацию о курсанте\n'
                f'<u>{help_course}</u>го курса\n'
                f'<u>{help_platoon}</u>го взвода\n'
                f'<u>{help_cadet.upper()}</u>?')
    elif help_platoon:
        info = (f'❓ Вы действительно хотите удалить информацию о\n'
                f'<u>{help_platoon}</u>м взводе\n'
                f'<u>{help_course}</u>го курса?')
    elif help_course:
        info = (f'❓ Вы действительно хотите удалить информацию о \n'
                f'<u>{help_course}</u>м курсе?')
    info += '\n\n❕ Отменить данное действие будет <u><b>НЕВОЗМОЖНО</b></u> ❕'

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                       one_time_keyboard=True,
                                       row_width=2,
                                       input_field_placeholder='Удалить информацию?')

    markup.add(types.KeyboardButton('🔚 Не удалять'), types.KeyboardButton('🗑 Удалить'))
    markup.add(types.KeyboardButton('🏠 На главную'))
    bot.send_message(chat_id=chat_id,
                     text=info,
                     reply_markup=markup,
                     parse_mode='html')
    bot.register_next_step_handler(message, delete_cadet_second, help_course=help_course, help_platoon=help_platoon,
                                   help_cadet=help_cadet)


def delete_cadet_second(message, help_course=None, help_platoon=None, help_cadet=None):
    chat_id = message.chat.id

    info = ''
    if help_cadet:
        info = (f'❓ Вы АБСОЛЮТНО УВЕРЕНЫ, что хотите удалить информацию о курсанте\n'
                f'<u>{help_course}</u>го курса\n'
                f'<u>{help_platoon}</u>го взвода\n'
                f'<u>{help_cadet.upper()}</u>?')
    elif help_platoon:
        info = (f'❓ Вы АБСОЛЮТНО УВЕРЕНЫ, что хотите удалить информацию о\n'
                f'<u>{help_platoon}</u>м взводе\n'
                f'<u>{help_course}</u>го курса?')
    elif help_course:
        info = (f'❓ Вы АБСОЛЮТНО УВЕРЕНЫ, что хотите удалить информацию о \n'
                f'<u>{help_course}</u>м курсе?')
    info += '\n\n❕ Отменить данное действие ТЕПЕРЬ будет <u><b>НЕВОЗМОЖНО</b></u> ❕'

    if message.text == '🗑 Удалить':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                           one_time_keyboard=True,
                                           row_width=2,
                                           input_field_placeholder='Вы уверены?')
        markup.add(types.KeyboardButton('🔚 НЕ УДАЛЯТЬ!'), types.KeyboardButton('🗑 Уверен, удалить!'))
        markup.add(types.KeyboardButton('🏠 На главную'))
        bot.send_message(chat_id=chat_id,
                         text=info,
                         reply_markup=markup,
                         parse_mode='html')
        bot.register_next_step_handler(message, apply_to_delete_cadet, help_course=help_course, help_platoon=help_platoon, help_cadet=help_cadet)
    elif message.text in command_list:
        return all_commands(message=message)
    elif message.text == '🔚 Не удалять':
        if help_platoon:
            return check_platoon(message, help_course=help_course, help_platoon=help_platoon, list_platoon=[], func_input_type='back', process_index=3)
        elif help_course:
            return check_course(message, help_course=help_course, list_kurs=[], func_input_type='back', process_index=3)
    elif message.text == '🏠 На главную':
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message)
    else:
        bot.send_message(chat_id=chat_id,
                         text='ℹ Нет такого варианта ответа!')
        return delete_cadet_first(message=message, help_course=help_course, help_platoon=help_platoon, help_cadet=help_cadet)


def apply_to_delete_cadet(message, help_course=None, help_platoon=None, help_cadet=None):
    chat_id = message.chat.id

    if message.text == '🗑 Уверен, удалить!':
        if help_platoon:
            with sqlite3.connect('./data_bases/students.sqlite3') as connection_to_db:
                cursor_db = connection_to_db.cursor()
            if help_cadet:
                cursor_db.execute('''
                    DELETE FROM 
                        parameters
                    WHERE 
                        student_id == 
                            (SELECT
                                student_id
                            FROM
                                cadets
                            WHERE
                                course == {course} AND platoon == {platoon} AND cadet == '{cadet}');
                '''
                                  .format(course=int(help_course),
                                          platoon=int(help_platoon),
                                          cadet=str(help_cadet)))

                cursor_db.execute('''
                    DELETE FROM 
                        cadets
                    WHERE 
                        course == {course} AND platoon == {platoon} AND cadet == '{cadet}';
                '''
                                  .format(course=int(help_course),
                                          platoon=int(help_platoon),
                                          cadet=str(help_cadet)))

                connection_to_db.commit()
                cursor_db.close()
                connection_to_db.close()

                bot.send_message(chat_id=chat_id,
                                 text=f'🚮 Данные о <u>{help_cadet.upper()}</u> удалены!',
                                 parse_mode='html')

                return check_platoon(message, help_course=help_course, help_platoon=help_platoon,
                                                        list_platoon=[], func_input_type='back', process_index=3)
            else:
                cursor_db.execute('''
                    DELETE FROM 
                        parameters
                    WHERE 
                        student_id == (
                            SELECT
                                student_id
                            FROM
                                cadets
                            WHERE
                                course == {course} AND platoon == {platoon});
                '''
                                  .format(course=int(help_course),
                                          platoon=int(help_platoon)))
                cursor_db.execute('''
                    DELETE FROM 
                        cadets 
                    WHERE 
                        course == ('%d') AND platoon == ('%d');
                '''
                                  % (int(help_course), int(help_platoon)))
                connection_to_db.commit()
                cursor_db.close()
                connection_to_db.close()

                bot.send_message(chat_id=chat_id,
                                 text=f'🚮 Данные о <u>{help_platoon}</u>м взводе удалены!',
                                 parse_mode='html')

                return check_course(message, help_course=help_course, list_kurs=[], func_input_type='back',
                                                    process_index=3)

        elif help_course:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                               one_time_keyboard=True,
                                               row_width=2,
                                               input_field_placeholder='Точно?')
            markup.add(types.KeyboardButton('🔚 Не удалять'), types.KeyboardButton('🗑 ТОЧНО!!!'))
            markup.add(types.KeyboardButton('🏠 На главную'))
            bot.send_message(chat_id=chat_id,
                             text=f'⁉ Вы <u><b>АБСОЛЮТНО ТОЧНО УВЕРЕНЫ</b></u>, что хотите <u><b>УДАЛИТЬ</b></u> информацию о \n'
                                  f'<u>{help_course}</u>м курсе?\n\n'
                                  f''
                                  f'‼ Отменить данное действие <u><b>ТЕПЕРЬ ТОЧНО</b></u> будет <u><b>НЕВОЗМОЖНО</b></u> ‼',
                             reply_markup=markup,
                             parse_mode='html')
            bot.register_next_step_handler(message, apply_to_delete_course, help_course=help_course)
    elif message.text in command_list:
        return all_commands(message=message)
    elif message.text == '🔚 НЕ УДАЛЯТЬ!':
        if help_platoon:
            return check_platoon(message, help_course=help_course, help_platoon=help_platoon, list_platoon=[], func_input_type='back', process_index=3)
        elif help_course:
            return check_course(message, help_course=help_course, list_kurs=[], func_input_type='back', process_index=3)
    elif message.text == '🏠 На главную':
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message)
    else:
        bot.send_message(chat_id=chat_id,
                         text='ℹ Нет такого варианта ответа!')
        bot.register_next_step_handler(message, apply_to_delete_cadet, help_course=help_course, help_platoon=help_platoon, help_cadet=help_cadet)


def apply_to_delete_course(message, help_course=None):
    chat_id = message.chat.id

    if message.text == '🗑 ТОЧНО!!!':
        with sqlite3.connect('./data_bases/students.sqlite3') as connection_to_db:
            cursor_db = connection_to_db.cursor()

        cursor_db.execute('''
            DELETE FROM 
                parameters
            WHERE 
                student_id == (
                    SELECT
                        student_id
                    FROM
                        cadets
                    WHERE
                        course == {course});
        '''
                          .format(course=int(help_course)))
        cursor_db.execute('''
            DELETE FROM 
                cadets 
            WHERE 
                course == ('%d');
        '''
                          % int(help_course))
        connection_to_db.commit()
        cursor_db.close()
        connection_to_db.close()

        bot.send_message(chat_id=chat_id,
                         text=f'🚮 Данные о <u>{help_course}</u>м курсе удалены!',
                         parse_mode='html')
        return command_all(message=message, do_edit_message=False, process_index=3)
    elif message.text in command_list:
        return all_commands(message=message)
    elif message.text == '🔚 Не удалять':
        return check_course(message, help_course=help_course, list_kurs=[], func_input_type='back', process_index=3)
    elif message.text == '🏠 На главную':
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message)
    else:
        bot.send_message(chat_id=chat_id,
                         text='ℹ Нет такого варианта ответа!')
        bot.register_next_step_handler(message, apply_to_delete_course, help_course=help_course)


def create_add_cadet_buttons(message, help_course=None, help_platoon=None, help_cadet=None):
    chat_id = message.chat.id

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                       one_time_keyboard=True,
                                       row_width=2,
                                       input_field_placeholder='Выберите нужный метод...')
    markup.add(types.KeyboardButton('🎓 Успеваемость'), types.KeyboardButton('🚨 Дисциплина'))
    markup.add(types.KeyboardButton('🔙 Назад'))
    markup.add(types.KeyboardButton('🏠 На главную'))
    bot.send_message(chat_id=chat_id,
                     text=f'Введите данные о <u><b>{str(help_cadet).upper()}</b></u>.',
                     reply_markup=markup,
                     parse_mode='html')
    bot.register_next_step_handler(message, add_info_about_cadet, help_course=help_course, help_platoon=help_platoon,
                                   help_cadet=help_cadet)


# занесение изменений в таблицу
def add_info_about_cadet(message, help_course=None, help_platoon=None, help_cadet=None):
    chat_id = message.chat.id
    with sqlite3.connect('./data_bases/students.sqlite3') as connection_to_db:
        cursor_db = connection_to_db.cursor()
    if (message.text == '🎓 Успеваемость') or (message.text == '🚨 Дисциплина'):
        if message.text == '🎓 Успеваемость':
            msg = bot.send_message(chat_id=chat_id,
                                   text='Добавляю информацию...')
            cursor_db.execute('''
                UPDATE 
                    parameters 
                SET 
                    grade = grade + 10, 
                    last_change = 'Успеваемость (+10)' 
                WHERE 
                    student_id == (
                        SELECT
                            student_id
                        FROM
                            cadets
                        WHERE
                            course == {course} AND platoon == {platoon} AND cadet == '{cadet}');
            '''
                              .format(course=int(help_course),
                                      platoon=int(help_platoon),
                                      cadet=str(help_cadet)))
            cursor_db.execute('''
                SELECT 
                    c.cadet, 
                    p.grade, 
                    p.discipline 
                FROM 
                    cadets AS c
                LEFT OUTER JOIN
                    parameters AS p
                ON
                    p.student_id = c.student_id
                WHERE 
                    c.course == {course} AND c.platoon == {platoon} AND c.cadet == '{cadet}';
            '''
                              .format(course=int(help_course),
                                      platoon=int(help_platoon),
                                      cadet=str(help_cadet)))
            users = cursor_db.fetchone()
            bot.edit_message_text(chat_id=chat_id,
                                  message_id=msg.message_id,
                                  text=f'Информация по <u>успеваемости</u> курсанта <u>{help_cadet}</u> <i>(+10)</i> добавлена!\n\n'
                                       f''
                                       f'<b>Имя</b>: <u>{users[0]}</u>\n'
                                       f'<b>Успеваемость</b>: <u>{users[1]}</u>\n'
                                       f'<b>Дисциплина</b>: <u>{users[2]}</u>',
                                  parse_mode='html')

        elif message.text == '🚨 Дисциплина':
            msg = bot.send_message(chat_id=chat_id,
                                   text='Добавляю информацию...')
            cursor_db.execute('''
                UPDATE 
                    parameters 
                SET 
                    discipline = discipline - 10, 
                    last_change = 'Дисциплина (-10)' 
                WHERE 
                    student_id == (
                        SELECT
                            student_id
                        FROM
                            cadets
                        WHERE
                            course == {course} AND platoon == {platoon} AND cadet == '{cadet}');
            '''
                              .format(course=int(help_course),
                                      platoon=int(help_platoon),
                                      cadet=str(help_cadet)))

            cursor_db.execute('''
                SELECT 
                    c.cadet, 
                    p.grade, 
                    p.discipline 
                FROM 
                    cadets AS c
                LEFT OUTER JOIN
                    parameters AS p
                ON
                    p.student_id = c.student_id
                WHERE 
                    c.course == {course} AND c.platoon == {platoon} AND c.cadet == '{cadet}';
            '''
                              .format(course=int(help_course),
                                      platoon=int(help_platoon),
                                      cadet=str(help_cadet)))
            users = cursor_db.fetchone()
            bot.edit_message_text(chat_id=chat_id,
                                  message_id=msg.message_id,
                                  text=f'Информация о <u>дисциплине</u> курсанта <u>{help_cadet}</u> <i>(-10)</i> добавлена!\n\n'
                                       f''
                                       f'<b>Имя</b>: <u>{users[0]}</u>\n'
                                       f'<b>Успеваемость</b>: <u>{users[1]}</u>\n'
                                       f'<b>Дисциплина</b>: <u>{users[2]}</u>',
                                  parse_mode='html')
        connection_to_db.commit()
        cursor_db.close()
        connection_to_db.close()

        return create_add_cadet_buttons(message=message, help_course=help_course, help_platoon=help_platoon, help_cadet=help_cadet)

    elif message.text == '🔙 Назад':
        return check_platoon(message, help_course=help_course, help_platoon=help_platoon, list_platoon=[],
                                                func_input_type='back', process_index=2)
    elif message.text == '🏠 На главную':
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message)
    elif message.text in command_list:
        return all_commands(message=message)
    else:
        bot.send_message(chat_id=chat_id,
                         text='ℹ Нет такого формата. Пожалуйста, выберите один из предложенных вариантов или нажмите кнопку \"🏠 <b>На главную</b>\" или \"🔙 <b>Назад</b>\".',
                         reply_markup=links_without_button_back(),
                         parse_mode='html')
        bot.register_next_step_handler(message, add_info_about_cadet, help_course=help_course, help_platoon=help_platoon,
                                       help_cadet=help_cadet)


# выбор метода получения информации в разделе "получение информации о пользователе"
def create_all_buttons(message, help_course=None, help_platoon=None, choice_info='faculty'):
    chat_id = message.chat.id

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                       one_time_keyboard=True,
                                       row_width=2,
                                       input_field_placeholder='Выберите нужный метод...')
    if choice_info == 'faculty':
        markup.add(types.KeyboardButton('💬 В чате (не рекомендуется)'), types.KeyboardButton('📃 В Excel таблице'))
    else:
        markup.add(types.KeyboardButton('💬 В чате'), types.KeyboardButton('📃 В Excel таблице'))
    markup.add(types.KeyboardButton('🔙 Назад'))
    markup.add(types.KeyboardButton('🏠 На главную'))
    bot.send_message(chat_id=chat_id,
                     text='❓ В каком формате Вам удобнее получить данные?',
                     reply_markup=markup)
    bot.register_next_step_handler(message, choice_all_method, help_course=help_course, help_platoon=help_platoon,
                                   choice_info=choice_info)


# получение информации о том, что выбрал пользователь в разделе "получение информации о пользователе"
def choice_all_method(message, help_course=None, help_platoon=None, choice_info='faculty'):
    chat_id = message.chat.id

    if message.text == '💬 В чате (не рекомендуется)' or message.text == '💬 В чате':
        return cadets_on_chat(message=message, help_course=help_course, help_platoon=help_platoon, choice_info=choice_info)
    if message.text == '📃 В Excel таблице':
        return cadets_on_xlsx(message=message, help_course=help_course, help_platoon=help_platoon, choice_info=choice_info)
    if message.text == '🔙 Назад' and choice_info == 'faculty':
        return command_all(message=message, do_edit_message=False, process_index=1)
    if message.text == '🔙 Назад' and choice_info == 'course':
        return check_course(message=message, help_course=help_course, list_kurs=[], func_input_type='back',
                                            process_index=1)
    if message.text == '🔙 Назад' and choice_info == 'platoon':
        return check_course(message=message, help_course=help_course, list_kurs=[], func_input_type='back',
                                            process_index=1)
    if message.text == '🏠 На главную':
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message)
    if message.text in command_list:
        return all_commands(message=message)

    bot.send_message(chat_id=chat_id,
                     text='ℹ Нет такого формата. Пожалуйста, выберите один из предложенных вариантов или нажмите кнопку \"🏠 <b>На главную</b>\" или \"🔙 <b>Назад</b>\".',
                     parse_mode='html')
    bot.register_next_step_handler(message, choice_all_method, help_course=help_course, help_platoon=help_platoon,
                                   choice_info=choice_info)


# отправка данных о деятельности курсанта в чат
def cadets_on_chat(message, help_course=None, help_platoon=None, choice_info='faculty'):
    chat_id = message.chat.id

    try:
        with sqlite3.connect('./data_bases/students.sqlite3') as connection_to_db:
            cursor_db = connection_to_db.cursor()

        if choice_info == 'faculty':
            cursor_db.execute('''
                SELECT 
                    c.course, 
                    c.platoon, 
                    c.cadet, 
                    p.grade, 
                    p.discipline 
                FROM 
                    cadets AS c
                LEFT OUTER JOIN
                    parameters AS p
                ON
                    c.student_id = p.student_id
                ORDER BY 
                    c.course, 
                    c.platoon, 
                    c.cadet;
            ''')
            faculty_info = cursor_db.fetchall()
            connection_to_db.commit()
            cursor_db.close()
            connection_to_db.close()

            info = ''
            for current_student in faculty_info:
                info += (f'<b>Курс</b>: {num_to_emoji(current_student[0])}, <b>Взвод</b>: {num_to_emoji(current_student[1])}\n'
                         f'<b>Имя</b>: <u>{current_student[2].upper()}</u>\n'
                         f'<b>Успеваемость</b>: {num_to_emoji(current_student[3])}, <b>Дисциплина</b>: {num_to_emoji(current_student[4])}\n\n')
            info = 'Информация о факультете:\n\n' + info
            info_to_chat(message=message, info=info, split_method='\n\n')

        elif choice_info == 'course':
            cursor_db.execute('''
                SELECT 
                    c.platoon, 
                    c.cadet, 
                    p.grade, 
                    p.discipline 
                FROM 
                    cadets AS c
                LEFT OUTER JOIN
                    parameters AS p
                ON
                    c.student_id = p.student_id
                WHERE
                    c.course == ('%d')
                ORDER BY 
                    c.platoon, 
                    c.cadet;
            '''
                              % (int(help_course)))
            course_info = cursor_db.fetchall()
            connection_to_db.commit()
            cursor_db.close()
            connection_to_db.close()

            info = ''
            for current_student in course_info:
                info += (f'<b>Взвод</b>: {num_to_emoji(current_student[0])}\n'
                         f'<b>Имя</b>: <u>{current_student[1].upper()}</u>\n'
                         f'<b>Успеваемость</b>: {num_to_emoji(current_student[2])}, <b>Дисциплина</b>: {num_to_emoji(current_student[3])}\n\n')

            info = f'Информация о {help_course}м курсе:\n\n' + info
            info_to_chat(message=message, info=info, split_method='\n\n')

        elif choice_info == 'platoon':
            cursor_db.execute('''
                SELECT 
                    c.cadet, 
                    p.grade, 
                    p.discipline 
                FROM 
                    cadets AS c
                LEFT OUTER JOIN
                    parameters AS p
                ON
                    c.student_id = p.student_id
                WHERE
                    c.platoon == ('%d')
                ORDER BY 
                    c.cadet;
            '''
                              % (int(help_platoon)))
            platoon_info = cursor_db.fetchall()
            connection_to_db.commit()
            cursor_db.close()
            connection_to_db.close()

            info = ''
            for current_student in platoon_info:
                info += (f'<b>Имя</b>: <u>{current_student[0].upper()}</u>\n'
                         f'<b>Успеваемость</b>: {num_to_emoji(current_student[1])}, <b>Дисциплина</b>: {num_to_emoji(current_student[2])}\n\n')

            info = f'Информация о {help_platoon}м взводе:\n\n' + info
            info_to_chat(message=message, info=info, split_method='\n\n')

        return create_all_buttons(message=message, help_course=help_course, help_platoon=help_platoon,
                                  choice_info=choice_info)

    except Exception as e:
        error_from_user(message=message, error_code=e)
        return create_all_buttons(message=message, help_course=help_course, help_platoon=help_platoon,
                                  choice_info=choice_info)


# отправка данных о деятельности курсанта в таблицу Excel
def cadets_on_xlsx(message, help_course=None, help_platoon=None, choice_info='faculty'):
    chat_id = message.chat.id

    with sqlite3.connect('./data_bases/students.sqlite3') as connection_to_db:
        cursor_db = connection_to_db.cursor()

    if choice_info == 'faculty':
        bot.reply_to(message, 'Создание \".xlsx\" файла...')
        cursor_db.execute('''
            SELECT 
                c.course, 
                c.platoon, 
                c.cadet, 
                p.grade, 
                p.discipline,
                p.last_change
            FROM 
                cadets AS c
            LEFT OUTER JOIN
                parameters AS p
            ON
                c.student_id = p.student_id
            ORDER BY 
                c.course, 
                c.platoon, 
                c.cadet;
        ''')
        cadets_in_faculty = cursor_db.fetchall()
        help_dictionary = {'Курс': {}, 'Взвод': {}, 'Курсант': {}, 'Успеваемость': {}, 'Дисциплина': {}, 'Последнее изменение': {}}
        count_string = 0
        for current_str_in_table in cadets_in_faculty:
            count_key = 0
            for key in help_dictionary.keys():
                help_dictionary[str(key)][str(count_string)] = current_str_in_table[int(count_key)]
                count_key += 1
            count_string += 1
        users_to_excel = pd.DataFrame(help_dictionary)
        users_to_excel.to_excel(f'Факультет.xlsx')
        bot.send_document(chat_id=chat_id,
                          document=open(f'Факультет.xlsx', 'rb'))
        os.remove(f'Факультет.xlsx')
    elif choice_info == 'course':
        bot.reply_to(message, 'Создание \".xlsx\" файла...')
        cursor_db.execute('''
            SELECT 
                c.platoon, 
                c.cadet, 
                p.grade, 
                p.discipline,
                p.last_change
            FROM 
                cadets AS c
            LEFT OUTER JOIN
                parameters AS p
            ON
                c.student_id = p.student_id
            WHERE
                c.course == ('%d')
            ORDER BY 
                c.platoon, 
                c.cadet;
        '''
                          % (int(help_course)))
        cadets_in_course = cursor_db.fetchall()
        help_dictionary = {'Взвод': {}, 'Курсант': {}, 'Успеваемость': {}, 'Дисциплина': {}, 'Последнее изменение': {}}
        count_string = 0
        for current_str_in_table in cadets_in_course:
            count_key = 0
            for key in help_dictionary.keys():
                help_dictionary[str(key)][str(count_string)] = current_str_in_table[int(count_key)]
                count_key += 1
            count_string += 1
        users_to_excel = pd.DataFrame(help_dictionary)
        users_to_excel.to_excel(f'{help_course}_курс.xlsx')
        bot.send_document(chat_id=chat_id,
                          document=open(f'{help_course}_курс.xlsx', 'rb'))
        os.remove(f'{help_course}_курс.xlsx')
    elif choice_info == 'platoon':
        bot.reply_to(message, 'Создание \".xlsx\" файла...')
        cursor_db.execute('''
            SELECT 
                c.cadet, 
                p.grade, 
                p.discipline,
                p.last_change
            FROM 
                cadets AS c
            LEFT OUTER JOIN
                parameters AS p
            ON
                c.student_id = p.student_id
            WHERE
                c.platoon == ('%d')
            ORDER BY 
                c.cadet;
        '''
                          % (int(help_platoon)))
        cadets_in_platoon = cursor_db.fetchall()
        help_dictionary = {'Курсант': {}, 'Успеваемость': {}, 'Дисциплина': {}, 'Последнее изменение': {}}
        count_string = 0
        for current_str_in_table in cadets_in_platoon:
            count_key = 0
            for key in help_dictionary.keys():
                help_dictionary[str(key)][str(count_string)] = current_str_in_table[int(count_key)]
                count_key += 1
            count_string += 1
        users_to_excel = pd.DataFrame(help_dictionary)
        users_to_excel.to_excel(f'{help_platoon}_взвод.xlsx')
        bot.send_document(chat_id=chat_id,
                          document=open(f'{help_platoon}_взвод.xlsx', 'rb'))
        os.remove(f'{help_platoon}_взвод.xlsx')

    connection_to_db.commit()
    cursor_db.close()
    connection_to_db.close()

    return create_all_buttons(message=message, help_course=help_course, help_platoon=help_platoon,
                              choice_info=choice_info)



# обновление данных таблицы
def update_info_every_year(message):
    chat_id = message.chat.id
    with sqlite3.connect('./data_bases/students.sqlite3') as connection_to_db:
        cursor_db = connection_to_db.cursor()

    bot.send_message(chat_id=chat_id,
                     text='❗❗ <b>УВЕДОМЛЕНИЕ!</b>\n'
                          '\t└ К началу учебного года <u>База Данных</u> с информацией о деятельности '
                          'курсантов <u>должна быть обновлена</u>, соответственно, высылаю Вам текущие '
                          'данные и обновляю Базу Данных.\n\n'
                          ''
                          'Рекомендуется проверить внесенные сейчас данные и, <u>если они не были '
                          'зарегистрированы, повторить их занесение</u>.',
                     parse_mode='html')
    cursor_db.execute('''
        SELECT 
            c.course, 
            c.platoon, 
            c.cadet, 
            p.grade, 
            p.discipline,
            p.last_change
        FROM 
            cadets AS c
        LEFT OUTER JOIN
            parameters AS p
        ON
            c.student_id = p.student_id
        ORDER BY 
            c.course, 
            c.platoon, 
            c.cadet;
    ''')
    students = cursor_db.fetchall()
    help_dictionary = {'Курс': {}, 'Взвод': {}, 'Курсант': {}, 'Успеваемость': {}, 'Дисциплина': {}, 'Последнее изменение': {}}
    count_string = 0
    for current_str_in_table in students:
        count_key = 0
        for key in help_dictionary.keys():
            help_dictionary[str(key)][str(count_string)] = current_str_in_table[int(count_key)]
            count_key += 1
        count_string += 1
    users_to_excel = pd.DataFrame(help_dictionary)
    users_to_excel.to_excel('Сводная_таблица_за_год_ФПСОИБ.xlsx')
    bot.send_document(message.chat.id, open('Сводная_таблица_за_год_ФПСОИБ.xlsx', 'rb'))
    os.remove('Сводная_таблица_за_год_ФПСОИБ.xlsx')

    cursor_db.execute('''
        UPDATE
            parameters
        SET
            grade = {value},
            discipline = {value};
    '''
                      .format(value=int(default_cadet_value)))

    connection_to_db.commit()
    cursor_db.close()
    connection_to_db.close()


def command_add_cadet(message, do_edit_message=False):
    chat_id = message.chat.id

    markup_inline = types.InlineKeyboardMarkup()
    markup_inline.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='from_mainpage_to_update_students'))
    markup_inline.add(types.InlineKeyboardButton(text='🏠 На главную', callback_data='back_to_main_page'))

    if do_edit_message:
        bot.delete_message(chat_id=chat_id,
                           message_id=bot.edit_message_text(chat_id=chat_id,
                                                            message_id=message.message_id,
                                                            text='⏳ Подготовка запроса...',
                                                            parse_mode='html').message_id)
        bot.send_message(chat_id=chat_id,
                         text=add_student_rules,
                         parse_mode='html',
                         reply_markup=markup_inline)

    else:
        bot.reply_to(message=message,
                     text=add_student_rules,
                     parse_mode='html',
                     reply_markup=markup_inline)

    bot.register_next_step_handler(message, check_users_info_about_cadet)


def check_users_info_about_cadet(message):
    chat_id = message.chat.id

    try:
        if message.text in command_list:
            return all_commands(message=message)

        if len(message.text.split('\n')) > len(message.text.split('|')):
            all_students = message.text.split('\n')
        else:
            all_students = message.text.split('|')

        info = ''

        for current_student in all_students:
            " ".join(current_student.split(' ')).strip()
            info_about_student = [x for x in current_student.split(' ') if x]
            if len(info_about_student) == 6:
                error_info = ''
                help_list = []

                if (len(info_about_student[0].lstrip('0')) == 1) and info_about_student[0].isdigit():
                    help_list.append(int(info_about_student[0].lstrip('0')))  # int
                else:
                    error_info += error_add_student[0]

                if ((len(info_about_student[1].lstrip('0')) >= 3) and (len(info_about_student[1].lstrip('0')) <= 4)
                        and (info_about_student[1][-1] != '0') and info_about_student[1].isdigit()):
                    help_list.append(int(info_about_student[1].lstrip('0')))  # int
                else:
                    error_info += error_add_student[1]

                if (len(info_about_student[2]) >= 2) and (len(info_about_student[3]) >= 2) and (
                        info_about_student[2][0] not in denied_symbols) and (
                        info_about_student[3][0] not in denied_symbols):
                    help_list.append(info_about_student[2] + ' ' + info_about_student[3])  # str
                else:
                    error_info += error_add_student[2]

                if info_about_student[4] == '_':
                    help_list.append(default_cadet_value)  # int
                elif info_about_student[4].isdigit():
                    if int(info_about_student[4]) in range(1, 3000):
                        help_list.append(int(info_about_student[4]))  # int
                else:
                    error_info += error_add_student[3]

                if info_about_student[5] == '_':
                    help_list.append(default_cadet_value)  # int
                elif info_about_student[5].isdigit():
                    if int(info_about_student[5]) in range(1, 3000):
                        help_list.append(int(info_about_student[5]))  # int
                else:
                    error_info += error_add_student[4]

                if len(help_list) == 5:
                    with sqlite3.connect('./data_bases/students.sqlite3') as connection_to_db:
                        cursor_db = connection_to_db.cursor()

                    cursor_db.execute('''
                        SELECT 
                            COUNT(*) 
                        FROM 
                            cadets 
                        WHERE 
                            course == ('%d') AND platoon == ('%d') AND cadet == ('%s');
                    '''
                                      % (help_list[0], help_list[1], help_list[2]))
                    do_have_from_db = cursor_db.fetchone()

                    do_have_from_data_set = False
                    for current_request in users[chat_id].data_set:
                        if ((current_request[0] == help_list[0]) and (current_request[1] == help_list[1]) and
                                (current_request[2] == help_list[2])):
                            do_have_from_data_set = True
                    if not do_have_from_data_set:
                        if not do_have_from_db[0]:
                            users[chat_id].data_set.append(help_list)
                            info += f'Данные о <u>{help_list[2]}</u> (<i>{help_list[0]}, {help_list[1]}</i>) могут быть добавлены!\n\n'
                        else:
                            info += f'Курсант <u>{help_list[0]}</u>го курса <u>{help_list[1]}</u>го взвода <u>{help_list[2]}</u> уже существует.\n\n'
                    else:
                        info += f'Вы уже указали в данном сообщении информацию о курсанте <u>{help_list[0]}</u>го курса <u>{help_list[1]}</u>го взвода <u>{help_list[2]}</u>.\n\n'

                    connection_to_db.commit()
                    cursor_db.close()
                    connection_to_db.close()

                else:
                    info += (
                        f'Неверно введены следующие поля (<u>{info_about_student[2]} {info_about_student[3]} ({info_about_student[0]}, {info_about_student[1]}</u>)):\n'
                        f'<b>{error_info}</b>\n')
            elif len(info_about_student) < 6:
                info += 'Заполнены не все поля!\n\n'
            else:
                info += 'Указана избыточная информация!\n\n'

        if users[chat_id].data_set:
            markup_inline = types.InlineKeyboardMarkup()
            markup_inline.add(types.InlineKeyboardButton(text='✔ Сохранить', callback_data='apply_to_save_add_student'))
            markup_inline.add(
                types.InlineKeyboardButton(text='❌ Не сохранять', callback_data='deny_to_save_add_student'))

            bot.send_message(chat_id=chat_id,
                             text=f'Вы хотите записать следующую информацию:\n\n'
                                  f''
                                  f'{info}'
                                  f''
                                  f'Сохранить изменения?',
                             parse_mode='html',
                             reply_markup=markup_inline)
        else:
            markup_inline = types.InlineKeyboardMarkup()
            markup_inline.add(
                types.InlineKeyboardButton(text='🔁 Повторить', callback_data='from_updatestudents_to_add_student'))
            markup_inline.add(types.InlineKeyboardButton(text='🏠 На главную', callback_data='back_to_main_page'))

            bot.send_message(chat_id=chat_id,
                             text=f'Невозможно записать следующую информацию:\n\n'
                                  f''
                                  f'{info}'
                                  f''
                                  f'Повторить отправку?',
                             parse_mode='html',
                             reply_markup=markup_inline)

    except AttributeError:
        markup_inline = types.InlineKeyboardMarkup()
        markup_inline.add(
            types.InlineKeyboardButton(text='🔁 Добавить информацию', callback_data='from_updatestudents_to_add_student'))
        markup_inline.add(types.InlineKeyboardButton(text='🏠 На главную', callback_data='back_to_main_page'))

        bot.send_message(chat_id=chat_id,
                         text='ℹ Невозможно определить действие. Хотите повторить отправку?',
                         parse_mode='html',
                         reply_markup=markup_inline)
    except telebot.apihelper.ApiTelegramException:
        error_from_user(message=message, error_code='Слишком длинный запрос! Отправьте, пожалуйста, необходимую информацию частями.')


def apply_add_student(message):
    chat_id = message.chat.id

    try:
        with sqlite3.connect('./data_bases/students.sqlite3') as connection_to_db:
            cursor_db = connection_to_db.cursor()
        for current_cadet in users[chat_id].data_set:
            cursor_db.execute('''
                INSERT INTO 
                    cadets (
                        course, 
                        platoon, 
                        cadet) 
                VALUES ('%d', '%d', '%s');
            '''
                              % (int(current_cadet[0]), int(current_cadet[1]), str(current_cadet[2])))
            cursor_db.execute('''
                INSERT INTO 
                    parameters (
                        student_id,
                        grade, 
                        discipline) 
                VALUES (
                    (SELECT
                        student_id
                    FROM
                        cadets
                    WHERE
                        course == ('%d') AND platoon == ('%d') AND cadet == ('%s')),
                    '%d', 
                    '%d');
            '''
                              % (int(current_cadet[0]), int(current_cadet[1]), str(current_cadet[2]), int(current_cadet[3]), int(current_cadet[4])))
        connection_to_db.commit()
        cursor_db.close()
        connection_to_db.close()

        users[chat_id].data_set = []

        markup_inline = types.InlineKeyboardMarkup()
        markup_inline.add(
            types.InlineKeyboardButton(text='🔁 Добавить информацию', callback_data='from_updatestudents_to_add_student'))
        markup_inline.add(types.InlineKeyboardButton(text='🏠 На главную', callback_data='back_to_main_page'))

        bot.edit_message_text(chat_id=chat_id,
                              message_id=message.message_id,
                              text='✔ Информация сохранена. Хотите добавить информацию о других курсантах?',
                              parse_mode='html',
                              reply_markup=markup_inline)
        users[chat_id].data_set.clear()

    except Exception as e:
        error_from_user(message=message, error_code=e)


def deny_add_student(message):
    chat_id = message.chat.id

    users[chat_id].data_set = []

    markup_inline = types.InlineKeyboardMarkup()
    markup_inline.add(
        types.InlineKeyboardButton(text='🔁 Добавить информацию', callback_data='from_updatestudents_to_add_student'))
    markup_inline.add(types.InlineKeyboardButton(text='🏠 На главную', callback_data='back_to_main_page'))

    bot.edit_message_text(chat_id=chat_id,
                          message_id=message.message_id,
                          text='❌ Отмена записи. Хотите добавить информацию о других курсантах?',
                          parse_mode='html',
                          reply_markup=markup_inline)
    users[chat_id].data_set.clear()


def command_update_user(message, process_index=1):
    chat_id = message.chat.id
    users[chat_id].update_user_list.clear()

    with sqlite3.connect('./data_bases/users.sqlite3') as connection_to_db:
        cursor_db = connection_to_db.cursor()
    cursor_db.execute('''
        SELECT
            user_login
        FROM
            users
    ''')
    all_users = cursor_db.fetchall()
    connection_to_db.commit()
    cursor_db.close()
    connection_to_db.close()

    help_list = []
    info = (f'👨‍💻🔁 В Базе Данных зарегистрировано {len(all_users)} пользователей:\n'
            f'\t│\n'
            f'')
    for current_user in all_users:
        help_list.append(str(current_user[0]))
        info += f'\t├ <code>{current_user[0]}</code>\n'
    info = info[:int(info.rfind('├'))] + '└' + info[int(info.rfind('├')) + 1:]
    info_to_chat(message=message, info=info, split_method='\n')

    if process_index == 1:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                           one_time_keyboard=True,
                                           row_width=2,
                                           input_field_placeholder='Выберите нужный логин...')
        markup.add(types.KeyboardButton('🔙 Назад'), types.KeyboardButton('🏠 На главную'))

        bot.send_message(chat_id=chat_id,
                         text=f'\n❔ Какого пользователя Вы хотите изменить?',
                         parse_mode='html',
                         reply_markup=markup)
        bot.register_next_step_handler(message, update_users_check_old_login, help_list=help_list)
    elif process_index == 2:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                           one_time_keyboard=True,
                                           row_width=2,
                                           input_field_placeholder='Выберите нужный логин...')
        markup.add(types.KeyboardButton('🔙 Назад'), types.KeyboardButton('🏠 На главную'))

        bot.send_message(chat_id=chat_id,
                         text=f'\n🗑 Какого пользователя Вы хотите удалить?',
                         parse_mode='html',
                         reply_markup=markup)
        bot.register_next_step_handler(message, delete_users_check_login, help_list=help_list)


def update_users_check_old_login(message, help_list=None, input_type=None):
    chat_id = message.chat.id

    if (message.text in help_list) or (input_type == 'back'):
        if input_type != 'back':
            users[chat_id].update_user_list.append(message.text)
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                           one_time_keyboard=True,
                                           row_width=2,
                                           input_field_placeholder='Введите новый логин...')
        markup.add(types.KeyboardButton('☑ Оставить без изменения'))
        markup.add(types.KeyboardButton('🔙 Назад'), types.KeyboardButton('🏠 На главную'))

        bot.send_message(chat_id=chat_id,
                         text=f'\n❔ Введите новый логин для пользователя <b>{users[chat_id].update_user_list[0]}</b>.',
                         parse_mode='html',
                         reply_markup=markup)
        bot.register_next_step_handler(message, update_users_check_new_login)
    elif message.text in command_list:
        return all_commands(message=message)
    elif message.text == '🔙 Назад':
        users[chat_id].update_user_list.clear()
        delete_reply_markup(message=message, text_hint='Возвращаю на страницу обновления пользователей...')
        return command_update_users(message=message, do_edit_message=False)
    elif message.text == '🏠 На главную':
        users[chat_id].update_user_list.clear()
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message=message, do_edit_message=False)
    else:
        bot.send_message(chat_id=chat_id,
                         text='ℹ Пользователя с данным логином не существует. Пожалуйста, введите логин из списка выше.')
        bot.register_next_step_handler(message, update_users_check_old_login, help_list=help_list)


def update_users_check_new_login(message, input_type=None):
    chat_id = message.chat.id

    if ((message.text) and (is_right_string(message.text))) or (input_type == 'back'):
        with sqlite3.connect('./data_bases/users.sqlite3') as connection_to_db:
            cursor_db = connection_to_db.cursor()
        if input_type != 'back':
            cursor_db.execute('''
                SELECT
                    COUNT(*)
                FROM
                    users
                WHERE
                    user_login == ('%s');
            '''
                              % (str(message.text)))
            count_users_for_current_login = cursor_db.fetchone()

            if count_users_for_current_login[0] != 0:
                connection_to_db.commit()
                cursor_db.close()
                connection_to_db.close()

                markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                                   one_time_keyboard=True,
                                                   row_width=2,
                                                   input_field_placeholder='Введите новый логин...')
                markup.add(types.KeyboardButton('☑ Оставить без изменения'))
                markup.add(types.KeyboardButton('🔙 Назад'), types.KeyboardButton('🏠 На главную'))

                bot.send_message(chat_id=chat_id,
                                 text=f'❕ Пользователь с данным логином уже существует.\n'
                                      f'Выберите, пожалуйста, другой логин.',
                                 parse_mode='html',
                                 reply_markup=markup)

                bot.register_next_step_handler(message, update_users_check_new_login)
            else:
                users[chat_id].update_user_list.append(message.text)
                cursor_db.execute('''
                    SELECT
                        user_password
                    FROM
                        users
                    WHERE
                        user_login == ('%s');
                '''
                                  % (str(users[chat_id].update_user_list[0])))
                current_user_password = cursor_db.fetchone()
                connection_to_db.commit()
                cursor_db.close()
                connection_to_db.close()

                markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                                   one_time_keyboard=True,
                                                   row_width=2,
                                                   input_field_placeholder='Введите новый пароль...')
                markup.add(types.KeyboardButton('☑ Оставить без изменения'))
                markup.add(types.KeyboardButton('🔙 Назад'), types.KeyboardButton('🏠 На главную'))

                bot.send_message(chat_id=chat_id,
                                 text=f'❕ <b>Старый пароль</b>:\n'
                                      f'\t└ <i>{current_user_password[0]}</i>\n\n'
                                      f''
                                      f'❔ Введите новый пароль для пользователя <b>{users[chat_id].update_user_list[0]}</b>.',
                                 parse_mode='html',
                                 reply_markup=markup)
                bot.register_next_step_handler(message, update_users_check_password)
        else:
            cursor_db.execute('''
                SELECT
                    user_password
                FROM
                    users
                WHERE
                    user_login == ('%s');
            '''
                              % (str(users[chat_id].update_user_list[0])))
            current_user_password = cursor_db.fetchone()
            connection_to_db.commit()
            cursor_db.close()
            connection_to_db.close()

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                               one_time_keyboard=True,
                                               row_width=2,
                                               input_field_placeholder='Введите новый пароль...')
            markup.add(types.KeyboardButton('☑ Оставить без изменения'))
            markup.add(types.KeyboardButton('🔙 Назад'), types.KeyboardButton('🏠 На главную'))

            bot.send_message(chat_id=chat_id,
                             text=f'❕ <b>Старый пароль</b>:\n'
                                  f'\t└ <i>{current_user_password[0]}</i>\n\n'
                                  f''
                                  f'❔ Введите новый пароль для пользователя <b>{users[chat_id].update_user_list[0]}</b>.',
                             parse_mode='html',
                             reply_markup=markup)
            bot.register_next_step_handler(message, update_users_check_password)
    elif message.text in command_list:
        users[chat_id].update_user_list.clear()
        return all_commands(message=message)
    elif message.text == '☑ Оставить без изменения':
        users[chat_id].update_user_list.append('*')
        return update_users_check_new_login(message, input_type='back')
    elif message.text == '🔙 Назад':
        users[chat_id].update_user_list.pop(-1)
        return command_update_user(message=message)
    elif message.text == '🏠 На главную':
        users[chat_id].update_user_list.clear()
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message=message, do_edit_message=False)
    else:
        bot.send_message(chat_id=chat_id,
                         text='ℹ Введен некорректный формат логина. Пожалуйста, повтороите ввод или воспользуйтесь навигационными кнопками ниже.')
        bot.register_next_step_handler(message, update_users_check_new_login)


def update_users_check_password(message, input_type=None):
    chat_id = message.chat.id

    if ((message.text) and (is_right_string(message.text))) or (input_type == 'back'):
        if input_type != 'back':
            users[chat_id].update_user_list.append(message.text)
        with sqlite3.connect('./data_bases/users.sqlite3') as connection_to_db:
            cursor_db = connection_to_db.cursor()
        cursor_db.execute('''
            SELECT
                user_role
            FROM
                users
            WHERE
                user_login == ('%s');
        '''
                          % (str(users[chat_id].update_user_list[0])))
        current_user_role = cursor_db.fetchone()
        connection_to_db.commit()
        cursor_db.close()
        connection_to_db.close()

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                           one_time_keyboard=True,
                                           row_width=4,
                                           input_field_placeholder='Введите новую роль...')
        markup.add(types.KeyboardButton('Гость'),
                   types.KeyboardButton('Курсант'),
                   types.KeyboardButton('Помощник'),
                   types.KeyboardButton('Офицер'))
        markup.add(types.KeyboardButton('☑ Оставить без изменения'))
        markup.add(types.KeyboardButton('🔙 Назад'), types.KeyboardButton('🏠 На главную'))

        bot.send_message(chat_id=chat_id,
                         text=f'❕ <b>Старая роль</b>:\n'
                              f'\t└ <i>{role_from_db_to_text(current_user_role[0], process_index=1)}</i>\n\n'
                              f''
                              f'❔ Введите новую роль для пользователя <b>{users[chat_id].update_user_list[0]}</b>.',
                         parse_mode='html',
                         reply_markup=markup)
        bot.register_next_step_handler(message, update_users_check_role)
    elif message.text in command_list:
        users[chat_id].update_user_list.clear()
        return all_commands(message=message)
    elif message.text == '☑ Оставить без изменения':
        users[chat_id].update_user_list.append('*')
        return update_users_check_password(message, input_type='back')
    elif message.text == '🔙 Назад':
        users[chat_id].update_user_list.pop(-1)
        return update_users_check_old_login(message, help_list=[], input_type='back')
    elif message.text == '🏠 На главную':
        users[chat_id].update_user_list.clear()
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message=message, do_edit_message=False)
    else:
        bot.send_message(chat_id=chat_id,
                         text='ℹ Введен некорректный формат пароля. Пожалуйста, повтороите ввод или воспользуйтесь навигационными кнопками ниже.')
        bot.register_next_step_handler(message, update_users_check_password)


def update_users_check_role(message, input_type=None):
    chat_id = message.chat.id

    if (message.text in role_list) or (input_type == 'back'):
        if input_type != 'back':
            users[chat_id].update_user_list.append(message.text)
        with sqlite3.connect('./data_bases/users.sqlite3') as connection_to_db:
            cursor_db = connection_to_db.cursor()
        cursor_db.execute('''
            SELECT
                tg_chat_id
            FROM
                users
            WHERE
                user_login == ('%s');
        '''
                          % (str(users[chat_id].update_user_list[0])))
        current_user_chat_id = cursor_db.fetchone()
        connection_to_db.commit()
        cursor_db.close()
        connection_to_db.close()

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                           one_time_keyboard=True,
                                           row_width=4,
                                           input_field_placeholder='Введите новый ID...')
        markup.add(types.KeyboardButton('☑ Оставить без изменения'))
        markup.add(types.KeyboardButton('🔙 Назад'), types.KeyboardButton('🏠 На главную'))

        bot.send_message(chat_id=chat_id,
                         text=f'❕ <b>Старый ID</b>:\n'
                              f'\t└ <i>{current_user_chat_id[0]}</i>\n\n'
                              f''
                              f'❔ Введите новый ID для пользователя <b>{users[chat_id].update_user_list[0]}</b>.',
                         parse_mode='html',
                         reply_markup=markup)
        bot.register_next_step_handler(message, update_users_check_chat_id)
    elif message.text in command_list:
        users[chat_id].update_user_list.clear()
        return all_commands(message=message)
    elif message.text == '☑ Оставить без изменения':
        users[chat_id].update_user_list.append('*')
        return update_users_check_role(message, input_type='back')
    elif message.text == '🔙 Назад':
        users[chat_id].update_user_list.pop(-1)
        return update_users_check_new_login(message, input_type='back')
    elif message.text == '🏠 На главную':
        users[chat_id].update_user_list.clear()
        delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
        return command_main_page(message=message, do_edit_message=False)
    else:
        bot.send_message(chat_id=chat_id,
                         text='ℹ Введен некорректный формат роли. Пожалуйста, повтороите ввод или воспользуйтесь навигационными кнопками ниже.')
        bot.register_next_step_handler(message, update_users_check_role)


def update_users_check_chat_id(message, input_type=None):
    chat_id = message.chat.id

    try:
        if (message.text).lstrip('0').isdigit() or input_type == 'back':
            if input_type != 'back':
                users[chat_id].update_user_list.append(message.text)
            with sqlite3.connect('./data_bases/users.sqlite3') as connection_to_db:
                cursor_db = connection_to_db.cursor()
            cursor_db.execute('''
                SELECT
                    user_login,
                    user_password,
                    user_role,
                    tg_chat_id
                FROM
                    users
                WHERE
                    user_login == ('%s');
            '''
                              % (str(users[chat_id].update_user_list[0])))
            current_user_info = cursor_db.fetchone()
            for num in range(1, len(users[chat_id].update_user_list)):
                if users[chat_id].update_user_list[num] == '*':
                    if num == 3:
                        users[chat_id].update_user_list[num] = role_from_db_to_text(str(current_user_info[num - 1]), process_index=1)
                    else:
                        users[chat_id].update_user_list[num] = str(current_user_info[num - 1])
            connection_to_db.commit()
            cursor_db.close()
            connection_to_db.close()

            info = (f'❕ Вы хотите обновить информацию о пользователе <b><u>{users[chat_id].update_user_list[0]}</u></b> в следующем виде:\n\n'
                    f''
                    f'Логин\n'
                    f'\t└ <u><b>{current_user_info[0]}</b></u> ➡ <u><b>{users[chat_id].update_user_list[1]}</b></u>\n'
                    f'Пароль\n'
                    f'\t└ <u><b>{current_user_info[1]}</b></u> ➡ <u><b>{users[chat_id].update_user_list[2]}</b></u>\n'
                    f'Роль\n'
                    f'\t└ <u><b>{role_from_db_to_text(current_user_info[2], process_index=1)}</b></u> ➡ <u><b>{users[chat_id].update_user_list[3]}</b></u>\n'
                    f'ID\n'
                    f'\t└ <u><b>{current_user_info[3]}</b></u> ➡ <u><b>{users[chat_id].update_user_list[4]}</b></u>\n\n'
                    f''
                    f'❔ Хотите обновить информацию о пользователе?')

            markup_inline = types.InlineKeyboardMarkup()
            markup_inline.row(types.InlineKeyboardButton(text='✔ Обновить', callback_data='apply_to_update_user'),
                              types.InlineKeyboardButton(text='❌ Отменить', callback_data='deny_to_update_user'))
            delete_reply_markup(message=message)
            bot.send_message(chat_id=chat_id,
                             text=info,
                             parse_mode='html',
                             reply_markup=markup_inline)
        elif message.text in command_list:
            users[chat_id].update_user_list.clear()
            return all_commands(message=message)
        elif message.text == '☑ Оставить без изменения':
            users[chat_id].update_user_list.append('*')
            return update_users_check_chat_id(message, input_type='back')
        elif message.text == '🔙 Назад':
            users[chat_id].update_user_list.pop(-1)
            return update_users_check_password(message=message, input_type='back')
        elif message.text == '🏠 На главную':
            users[chat_id].update_user_list.clear()
            delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
            return command_main_page(message=message, do_edit_message=False)
        else:
            bot.send_message(chat_id=chat_id,
                             text='ℹ Введен некорректный формат ID. Пожалуйста, повтороите ввод или воспользуйтесь навигационными кнопками ниже.')
            bot.register_next_step_handler(message, update_users_check_chat_id)
    except AttributeError:
        bot.send_message(chat_id=chat_id,
                         text='ℹ Введен некорректный формат ID. Пожалуйста, повтороите ввод или воспользуйтесь навигационными кнопками ниже.')
        bot.register_next_step_handler(message, update_users_check_chat_id)


def apply_to_update_user(message):
    chat_id = message.chat.id

    try:
        with sqlite3.connect('./data_bases/users.sqlite3') as connection_to_db:
            cursor_db = connection_to_db.cursor()
        cursor_db.execute('''
            UPDATE
                users
            SET
                user_login = ('%s'),
                user_password = ('%s'),
                user_role = ('%s'),
                tg_chat_id = ('%d')
            WHERE
                user_login == ('%s');
        '''
                          % (str(users[chat_id].update_user_list[1]),
                             str(users[chat_id].update_user_list[2]),
                             role_from_db_to_text(str(users[chat_id].update_user_list[3]), process_index=2),
                             int(users[chat_id].update_user_list[4]),
                             str(users[chat_id].update_user_list[0])))
        connection_to_db.commit()
        cursor_db.close()
        connection_to_db.close()

        markup_inline = types.InlineKeyboardMarkup()
        markup_inline.add(types.InlineKeyboardButton(text='🔁 Повторить', callback_data='from_updateusers_to_update_user'))
        markup_inline.add(types.InlineKeyboardButton(text='🏠 На главную', callback_data='back_to_main_page'))

        bot.edit_message_text(chat_id=chat_id,
                              message_id=message.message_id,
                              text=f'✔ Данные о пользователе <u>{users[chat_id].update_user_list[0]}</u> обновлены!\n\n'
                                   f''
                                   f'❔ Хотите повторить обновление для другого пользователя?',
                              parse_mode='html',
                              reply_markup=markup_inline)
        users[chat_id].update_user_list.clear()
    except Exception as e:
        return error_from_user(message=message, error_code=e)


def deny_to_update_user(message):
    chat_id = message.chat.id

    try:
        markup_inline = types.InlineKeyboardMarkup()
        markup_inline.add(types.InlineKeyboardButton(text='🔁 Повторить', callback_data='from_updateusers_to_update_user'))
        markup_inline.add(types.InlineKeyboardButton(text='🏠 На главную', callback_data='back_to_main_page'))

        bot.edit_message_text(chat_id=chat_id,
                              message_id=message.message_id,
                              text=f'❌ Отмена обновления данных о пользователе <u>{users[chat_id].update_user_list[0]}</u>.\n\n'
                                   f''
                                   f'❔ Хотите повторить обновление для другого пользователя?',
                              parse_mode='html',
                              reply_markup=markup_inline)
        users[chat_id].update_user_list.clear()
    except Exception as e:
        return error_from_user(message=message, error_code=e)


def delete_users_check_login(message, help_list=None):
    chat_id = message.chat.id
    try:
        if message.text in help_list:
            users[chat_id].user_to_delete = message.text
            markup_inline = types.InlineKeyboardMarkup()
            markup_inline.row(types.InlineKeyboardButton(text='🗑 Удалить', callback_data='apply_to_delete_user'),
                              types.InlineKeyboardButton(text='❌ Не удалять', callback_data='deny_to_delete_user'))
            delete_reply_markup(message=message)
            bot.send_message(chat_id=chat_id,
                             text=f'❓ Вы действительно хотите удалить пользователя {users[chat_id].user_to_delete}',
                             parse_mode='html',
                             reply_markup=markup_inline)

        elif message.text in command_list:
            return all_commands(message=message)
        elif message.text == '🔙 Назад':
            delete_reply_markup(message=message, text_hint='Возвращаю на страницу обновления пользователей...')
            return command_update_users(message=message, do_edit_message=False)
        elif message.text == '🏠 На главную':
            delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
            return command_main_page(message=message, do_edit_message=False)
        else:
            bot.send_message(chat_id=chat_id,
                             text='ℹ Пользователя с данным логином не существует. Пожалуйста, введите логин из списка выше.')
            bot.register_next_step_handler(message, delete_users_check_login, help_list=help_list)
    except Exception as e:
        return error_from_user(message=message, error_code=e)


def apply_to_delete_user(message):
    chat_id = message.chat.id

    try:
        with sqlite3.connect('./data_bases/users.sqlite3') as connection_to_db:
            cursor_db = connection_to_db.cursor()
        cursor_db.execute('''
            DELETE FROM
                users
            WHERE
                user_login == ('%s');
        '''
                          % (str(users[chat_id].user_to_delete)))
        connection_to_db.commit()
        cursor_db.close()
        connection_to_db.close()

        markup_inline = types.InlineKeyboardMarkup()
        markup_inline.row(types.InlineKeyboardButton(text='🔙 Назад', callback_data='from_updateusers_to_delete_user'),
                          types.InlineKeyboardButton(text='🏠 На главную', callback_data='back_to_main_page'))
        bot.edit_message_text(chat_id=chat_id,
                              message_id=message.message_id,
                              text=f'Пользователь с логином <u>{users[chat_id].user_to_delete}</u> успешно удален!',
                              parse_mode='html',
                              reply_markup=markup_inline)
        users[chat_id].user_to_delete = ''
    except Exception as e:
        return error_from_user(message=message, error_code=e)


def deny_to_delete_user(message):
    chat_id = message.chat.id

    markup_inline = types.InlineKeyboardMarkup()
    markup_inline.row(types.InlineKeyboardButton(text='🔙 Назад', callback_data='from_updateusers_to_delete_user'),
                      types.InlineKeyboardButton(text='🏠 На главную', callback_data='back_to_main_page'))
    bot.edit_message_text(chat_id=chat_id,
                          message_id=message.message_id,
                          text=f'❌ Отмена удаления пользователя с логином <u>{users[chat_id].user_to_delete}</u>.',
                          parse_mode='html',
                          reply_markup=markup_inline)
    users[chat_id].user_to_delete = ''


@bot.message_handler()
def some_text_from_user(message):
    chat_id = message.chat.id

    try:
        global start_year_of_start
        global last_num_year_of_start
        now_data = date.today()
        if ((int(now_data.year) > start_year_of_start) and (int(now_data.month) >= 8) and (
                int(now_data.day) >= 10)):  # если нынешний год > год начала работы программы И дата >= 10 августа
            start_year_of_start = int(now_data.year)
            last_num_year_of_start = start_year_of_start % 10
            update_info_every_year(message)
            bot.send_message(chat_id=chat_id,
                             text='🔁 База Данных курсантов обновлена!')

        if users[chat_id].user_role != 'guest':
            if message.text == '🏠 На главную':
                delete_reply_markup(message=message, text_hint='Возвращаю на главную страницу...')
                return command_main_page(message)
            else:
                bot.send_message(chat_id=chat_id,
                                 text='ℹ Невозможно определить действие по заданным параметрам.')

        else:
            markup_inline = types.InlineKeyboardMarkup()
            markup_inline.add(types.InlineKeyboardButton(text='▶ Да, зарегистрироваться', callback_data='signup'),
                              types.InlineKeyboardButton(text='❌ Нет, остаться', callback_data='back_to_start'))
            bot.send_message(chat_id=chat_id,
                             text=f'❕ Вы не зарегистрированы! Зарегистрироваться? ❕',
                             reply_markup=markup_inline)

    except KeyError:
        delete_reply_markup(message=message,
                            text_hint='Устранение конфликтов...')
        users[chat_id] = User()
        users[chat_id].user_id = message.from_user.id
        users[chat_id].user_name = message.from_user.first_name
        users[chat_id].user_role = 'guest'

        bot.send_message(chat_id=chat_id,
                         text=f'🛠 Теперь Вам доступна работа с ботом!')


@bot.callback_query_handler(func=lambda callback: True)
def callback_message(callback):
    chat_id = callback.message.chat.id

    # если пользователь нажал на кнопку, то все степ хэндлеры прекращаются
    bot.clear_step_handler_by_chat_id(chat_id=chat_id)

    try:
        if users[chat_id].user_role:
            if callback.data == 'from_mainpage_to_help':
                return command_help(callback.message, do_edit_message=True)
            elif callback.data == 'from_mainpage_to_info_about_user':
                return command_my_id(callback.message, do_edit_message=True)

            elif users[chat_id].user_role == 'guest':
                if callback.data == 'login':
                    return login_in_system(callback.message, do_edit_message=True)
                elif callback.data == 'signup':
                    return signup_in_system(callback.message)
                elif callback.data == 'input_other_login':
                    return login_in_system(callback.message, do_edit_message=True)
                elif callback.data == 'back_to_start':
                    return command_start(callback.message, do_edit_message=True)

            # для всех, кроме гостей
            elif users[chat_id].user_role != 'guest':
                if callback.data == 'back_to_main_page':
                    return command_main_page(callback.message, do_edit_message=True)

                elif callback.data == 'stay_on_system':
                    return command_main_page(callback.message, do_edit_message=True)
                elif callback.data == 'log_out':
                    return logout_function(message=callback.message)

                elif callback.data == 'from_mainpage_to_schedule':
                    return command_schedule(callback.message, do_edit_message=True)
                elif callback.data == 'from_mainpage_to_all':
                    bot.answer_callback_query(callback_query_id=callback.id,
                                              show_alert=False,
                                              text='Выбрано: 👥 Информация о курсантах')
                    return command_all(callback.message, do_edit_message=True, process_index=1)

                elif callback.data == 'today_schedule_callback':
                    bot.answer_callback_query(callback_query_id=callback.id,
                                              show_alert=False,
                                              text='Выбрано: 📆 Текущее расписание')
                    return create_schedule_courses(callback.message, do_edit_message=True, process_index=1)
                elif callback.data == 'subscribe_to_schedule_callback':
                    bot.answer_callback_query(callback_query_id=callback.id,
                                              show_alert=False,
                                              text='Выбрано: 🔔 Подписаться на уведомление о расписании')
                    return create_schedule_courses(callback.message, do_edit_message=True, process_index=2)
                elif callback.data == 'unsubscribe_to_schedule_callback':
                    bot.answer_callback_query(callback_query_id=callback.id,
                                              show_alert=False,
                                              text='Выбрано: 🚫 Отменить подписку на уведомление о расписании')
                    return unsubscribe_to_schedule(callback.message)

                elif callback.data == 'back_to_schedule':
                    return command_schedule(callback.message, do_edit_message=True)

                elif users[chat_id].user_role != 'cadet':
                    if callback.data == 'from_mainpage_to_add':
                        bot.answer_callback_query(callback_query_id=callback.id,
                                                  show_alert=False,
                                                  text='Выбрано: ➕ Добавить информацию о курсантах')
                        return command_all(callback.message, do_edit_message=True, process_index=2)

                    elif callback.data == 'add_file_from_user_callback':
                        bot.answer_callback_query(callback_query_id=callback.id,
                                                  show_alert=False,
                                                  text='Выбрано: ⬇ Добавить файл расписания курса/взвода')
                        return schedule_document_rules(message=callback.message)
                    elif callback.data == 'apply_to_save_document':
                        return apply_to_save_schedule_document(message=callback.message)
                    elif callback.data == 'deny_to_save_document':
                        return deny_to_save_schedule_document(message=callback.message)

                    elif users[chat_id].user_role != 'helper':
                        if callback.data == 'from_mainpage_to_update_students':
                            return update_students(callback.message, do_edit_message=True)
                        elif callback.data == 'from_updatestudents_to_add_student':
                            bot.answer_callback_query(callback_query_id=callback.id,
                                                      show_alert=False,
                                                      text='Выбрано: 👨🎓 Добавить курсанта')
                            return command_add_cadet(callback.message, do_edit_message=True)
                        elif callback.data == 'from_updatestudents_to_update_student':
                            bot.answer_callback_query(callback_query_id=callback.id,
                                                      show_alert=False,
                                                      text='Выбрано: 👨🔁 Обновить курсанта')
                            return command_all(callback.message, do_edit_message=True, process_index=4)
                        elif callback.data == 'from_updatestudents_to_delete_student':
                            bot.answer_callback_query(callback_query_id=callback.id,
                                                      show_alert=False,
                                                      text='Выбрано: 👨❌ Удалить курсанта')
                            return command_all(callback.message, do_edit_message=True, process_index=3)
                        elif callback.data == 'apply_to_save_add_student':
                            return apply_add_student(message=callback.message)
                        elif callback.data == 'deny_to_save_add_student':
                            return deny_add_student(message=callback.message)

                        elif callback.data == 'from_mainpage_to_update_users':
                            return command_update_users(callback.message, do_edit_message=True)
                        elif callback.data == 'from_updateusers_to_update_user':
                            bot.answer_callback_query(callback_query_id=callback.id,
                                                      show_alert=False,
                                                      text='Выбрано: 👨‍💻🔁 Обновить пользователя')
                            return command_update_user(callback.message, process_index=1)
                        elif callback.data == 'apply_to_update_user':
                            return apply_to_update_user(message=callback.message)
                        elif callback.data == 'deny_to_update_user':
                            return deny_to_update_user(message=callback.message)

                        elif callback.data == 'from_updateusers_to_delete_user':
                            bot.answer_callback_query(callback_query_id=callback.id,
                                                      show_alert=False,
                                                      text='Выбрано: 👨‍💻❌ Удалить пользователя')
                            return command_update_user(callback.message, process_index=2)
                        elif callback.data == 'apply_to_delete_user':
                            return apply_to_delete_user(message=callback.message)
                        elif callback.data == 'deny_to_delete_user':
                            return deny_to_delete_user(message=callback.message)

        bot.answer_callback_query(callback_query_id=callback.id,
                                  show_alert=True,
                                  text='Вы не имеете прав на доступ к данным функциям!')
        bot.delete_message(chat_id=chat_id,
                           message_id=callback.message.message_id)
        return

    except KeyError:
        users[chat_id] = User()
        users[chat_id].user_role = 'guest'
        bot.answer_callback_query(callback_query_id=callback.id,
                                  show_alert=True,
                                  text='Вы не имеете прав на доступ к данным функциям!')
        bot.delete_message(chat_id=chat_id,
                           message_id=callback.message.message_id)


# принятие документа расписания
def take_document(message):
    chat_id = message.chat.id

    msg = bot.reply_to(message, text="Присходит анализ...")

    try:
        if message.document.file_name.split('.')[-1] == 'xlsx':
            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)

            users[chat_id].excel_schedule_file_name = './xlsx_files/' + message.document.file_name
            with open(users[chat_id].excel_schedule_file_name, 'wb') as new_file:
                new_file.write(downloaded_file)

            delete_reply_markup(message=message, text_hint='Следующий шаг...')
            users[chat_id].groups_in_xlsx_save = check_schedule_document_from_user(message, msg_help=msg)
        else:
            bot.delete_message(chat_id=chat_id,
                               message_id=msg.message_id)
            return schedule_document_rules_error(message)
    except AttributeError:
        bot.delete_message(chat_id=chat_id,
                           message_id=msg.message_id)
        if message.text == '🔙 Назад':
            delete_reply_markup(message=message, text_hint='Возвращаю на страницу расписания...')
            return command_schedule(message=message, do_edit_message=False)
        elif message.text in command_list:
            return all_commands(message=message)
        else:
            return schedule_document_rules_error(message=message)


@bot.message_handler(content_types=['document'])
def error_send_file(message):
    chat_id = message.chat.id

    try:
        if (users[chat_id].user_role == 'officer') or (users[chat_id].user_role == 'helper'):
            bot.reply_to(message=message,
                         text='Пожалуйста, сначала выберите соответствующий раздел в меню, а затем уже отправьте нужный файл.')
        else:
            bot.reply_to(message=message,
                         text='Вы не имеете прав на отправку боту файлов!')

    except KeyError:
        users[chat_id] = User()
        users[chat_id].user_id = message.from_user.id
        users[chat_id].user_name = message.from_user.first_name
        users[chat_id].user_role = 'guest'
        bot.reply_to(message=message,
                     text='У Вас недостаточно прав!')


if __name__ == '__main__':
    bot.polling(none_stop=True)
